import argparse
import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

EXCEL_EXTS = {'.xls', '.xlsx', '.xlsm'}
INPUT_EXTS = EXCEL_EXTS | {'.lnk'}
LEFT_BLOCK_COLS = ['A', 'B', 'C', 'D', 'E']
RIGHT_BLOCK_COLS = ['G', 'H', 'I', 'J', 'K']
TOLERANCE = 0.01


def log(msg):
    print(msg)


def strip_quotes(s):
    return str(s or '').strip().strip('"').strip("'")


def normalize_award(value):
    if value is None:
        return ''
    s = str(value).strip().upper()
    m = re.search(r'([A-Z0-9]{5})', s)
    return m.group(1) if m else s


def normalize_task(value):
    if value is None:
        return ''
    s = str(value).strip()
    if s.endswith('.0'):
        s = s[:-2]
    m = re.search(r'(\d{3,5})', s)
    return m.group(1) if m else s


def parse_month(month_text):
    raw = strip_quotes(month_text).lower()
    for fmt in ('%b-%Y', '%B-%Y', '%b-%y', '%B-%y', '%Y-%m'):
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.year, dt.month, dt.strftime('%b-%y').lower(), dt.strftime('%b-%Y').upper()
        except ValueError:
            pass
    raise ValueError(f"Could not parse month '{month_text}'. Try forms like mar-2026 or March-2026.")


def safe_float(value):
    try:
        if value is None or value == '':
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def extract_code(long_name):
    s = str(long_name or '').strip()
    m = re.match(r'(\d{5})', s)
    return m.group(1) if m else s[:5]


def parse_award_task_from_name(name):
    stem = Path(name).stem.upper()
    # Examples: UBEDQ.601.xlsx, UBEDQ - 601.xlsm, PCQAN_601.xlsx
    matches = list(re.finditer(r'([A-Z0-9]{5})[^A-Z0-9]+(\d{3,5})(?!\d)', stem))
    if matches:
        last = matches[-1]
        return last.group(1), last.group(2)
    tokens = re.findall(r'[A-Z0-9]+', stem)
    for i in range(len(tokens) - 1, 0, -1):
        if re.fullmatch(r'\d{3,5}', tokens[i]) and re.fullmatch(r'[A-Z0-9]{5}', tokens[i - 1]):
            return tokens[i - 1], tokens[i]
    return None, None


def parse_award_sum_from_name(name):
    stem = Path(name).stem.upper()
    # Examples: UBEDQ.SUM.xlsx, UBEDQ SUM.xlsm
    m = re.search(r'([A-Z0-9]{5})[^A-Z0-9]*SUM\b', stem)
    if m:
        return m.group(1)
    tokens = re.findall(r'[A-Z0-9]+', stem)
    if tokens and tokens[-1] == 'SUM':
        for tok in reversed(tokens[:-1]):
            if re.fullmatch(r'[A-Z0-9]{5}', tok):
                return tok
    return None


def is_sum_file(path):
    return parse_award_sum_from_name(Path(path).name) is not None


def ensure_dir(path):
    if path and not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def resolve_shortcut(path):
    try:
        import win32com.client
    except Exception as exc:
        raise RuntimeError('pywin32 is required to resolve .lnk shortcuts on Windows.') from exc
    shell = win32com.client.Dispatch('WScript.Shell')
    shortcut = shell.CreateShortcut(path)
    target = shortcut.TargetPath
    if not target:
        raise RuntimeError(f'Could not resolve shortcut target for {path}')
    return target


def iter_input_files(inputs, include_subfolders=False):
    seen = set()
    files = []
    for item in inputs:
        item = strip_quotes(item)
        if not item:
            continue
        p = Path(item)
        if p.is_dir():
            iterator = p.rglob('*') if include_subfolders else p.glob('*')
            for child in iterator:
                if child.is_file() and child.suffix.lower() in INPUT_EXTS:
                    if child.name.startswith('~$'):
                        continue
                    if child.name.lower() in {'commit_fill_report.xlsx', 'commit_fill_report.csv'}:
                        continue
                    key = str(child.resolve()) if child.exists() else str(child)
                    if key not in seen:
                        seen.add(key)
                        files.append(str(child))
        elif p.is_file() and p.suffix.lower() in INPUT_EXTS:
            if p.name.startswith('~$'):
                continue
            key = str(p.resolve()) if p.exists() else str(p)
            if key not in seen:
                seen.add(key)
                files.append(str(p))
    return files


def clean_column_name(value):
    s = str(value or '').strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def column_lookup_name(value):
    # Pandas adds .1, .2, etc. for duplicate Excel headers.
    # This lets duplicate columns still match the original header text.
    s = clean_column_name(value).lower()
    s = re.sub(r'\.\d+$', '', s)
    s = re.sub(r'__\d+$', '', s)
    return s


def make_unique_columns(columns):
    seen = {}
    result = []
    for col in columns:
        base = clean_column_name(col)
        if base not in seen:
            seen[base] = 1
            result.append(base)
        else:
            seen[base] += 1
            result.append(f'{base}__{seen[base]}')
    return result


def find_report_columns(columns):
    cleaned = {}
    for c in columns:
        key = column_lookup_name(c)
        # Keep the first duplicate, matching the behavior of the older script.
        if key not in cleaned:
            cleaned[key] = c
    aliases = {
        'Task Number': ['task number', 'task', 'task no', 'task id'],
        'Award Number': ['award number', 'award', 'award no', 'award id'],
        'Expenditure Type Long Name': ['expenditure type long name', 'expense type long name', 'expenditure type', 'expenditure type name'],
        'Purchase Order Number': ['purchase order number', 'po number', 'purchase order', 'po'],
        'Expenditure Item Date': ['expenditure item date', 'item date', 'expenditure date', 'date'],
        'Expenditure Commitment Amount': ['expenditure commitment amount', 'commitment amount', 'commit amount', 'po commitment amount', 'amount'],
    }
    mapping = {}
    missing = []
    for standard, names in aliases.items():
        found = None
        for name in names:
            if name in cleaned:
                found = cleaned[name]
                break
        if found is None:
            words = re.findall(r'[a-z0-9]+', standard.lower())
            for actual in columns:
                actual_clean = column_lookup_name(actual)
                if all(w in actual_clean for w in words):
                    found = actual
                    break
        if found is None:
            missing.append(standard)
        else:
            mapping[standard] = found
    return mapping, missing


def excel_serial_to_datetime(value):
    if pd.isna(value):
        return pd.NaT
    if isinstance(value, (datetime, pd.Timestamp)):
        return pd.to_datetime(value, errors='coerce')
    try:
        if isinstance(value, (int, float)) and 20000 < float(value) < 90000:
            return pd.to_datetime(float(value), unit='D', origin='1899-12-30', errors='coerce')
    except Exception:
        pass
    return pd.to_datetime(value, errors='coerce')


def is_indirect_row(row):
    high = str(row.get('High Level', '') or '').upper()
    type_long = str(row.get('Type Long', '') or '').upper()
    code = str(row.get('CODE', '') or '').strip()
    return (
        'INDIRECT' in high
        or code == '56910'
        or 'FACILITIES' in type_long
        or 'ADMIN CHARGE' in type_long
    )


def is_equipment_row(row):
    high = str(row.get('High Level', '') or '').upper()
    type_long = str(row.get('Type Long', '') or '').upper()
    code = str(row.get('CODE', '') or '').strip()
    return (
        'EQUIPMENT' in high
        or 'DEPRECIATION' in high
        or code == '53115'
        or 'CAP SCIENTIFIC TECH EQUIP' in type_long
    )


def read_commitment_report(report_path):
    log(f"[STEP] Reading raw commitment report: {report_path}")

    xl = pd.ExcelFile(report_path)
    raw = None
    used_sheet = None
    used_header = None
    used_mapping = None
    best_guess = None

    for sheet_name in xl.sheet_names:
        for header_row in range(0, 31):
            try:
                candidate = pd.read_excel(report_path, sheet_name=sheet_name, header=header_row)
            except Exception:
                continue
            candidate.columns = make_unique_columns(candidate.columns)
            mapping, missing = find_report_columns(candidate.columns)
            found_count = 6 - len(missing)
            if best_guess is None or found_count > best_guess['found_count']:
                best_guess = {
                    'sheet': sheet_name,
                    'header_row': header_row,
                    'found_count': found_count,
                    'missing': missing,
                    'columns': [str(c) for c in candidate.columns[:60]],
                }
            if not missing:
                raw = candidate
                used_sheet = sheet_name
                used_header = header_row
                used_mapping = mapping
                break
        if raw is not None:
            break

    if raw is None:
        if best_guess:
            raise KeyError(
                "Raw commitment report columns were not found automatically. "
                f"Best guess: sheet '{best_guess['sheet']}', header row {best_guess['header_row'] + 1}, "
                f"found {best_guess['found_count']}/6 columns. Missing: {', '.join(best_guess['missing'])}. "
                f"Columns seen: {best_guess['columns']}"
            )
        raise KeyError(f"Raw commitment report columns were not found automatically. Tried sheets {xl.sheet_names}.")

    log(f"[STEP] Using sheet '{used_sheet}' with header row {used_header + 1}")

    rename_map = {actual: standard for standard, actual in used_mapping.items()}
    raw = raw.rename(columns=rename_map).copy()

    first_col = raw.columns[0]
    raw = raw[~raw[first_col].astype(str).str.strip().isin(['Grand Total', 'Rows 1 - 90 (All Rows)'])].copy()

    # Some Oracle-style exports use blank merged-looking cells. Forward-fill only
    # identity/category fields, never amount/date/PO.
    ffill_cols = [
        'Project Number', 'Task Number', 'Award Number',
        'Expenditure High Level Group Description', 'Expenditure Type Long Name',
        'Source System Description'
    ]
    for col in ffill_cols:
        if col in raw.columns:
            raw[col] = raw[col].ffill()

    raw['Award Number'] = raw['Award Number'].map(normalize_award)
    raw['Task Number'] = raw['Task Number'].map(normalize_task)
    raw['Purchase Order Number'] = raw['Purchase Order Number'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    raw['Expenditure Item Date'] = raw['Expenditure Item Date'].map(excel_serial_to_datetime)
    raw['Expenditure Commitment Amount'] = pd.to_numeric(raw['Expenditure Commitment Amount'], errors='coerce').fillna(0.0)
    raw['CODE'] = raw['Expenditure Type Long Name'].map(extract_code)

    if 'Expenditure High Level Group Description' in raw.columns:
        raw['High Level'] = raw['Expenditure High Level Group Description'].fillna('').astype(str)
    else:
        raw['High Level'] = ''

    if 'Expenditure Type Long Name' in raw.columns:
        raw['Type Long'] = raw['Expenditure Type Long Name'].fillna('').astype(str)
    else:
        raw['Type Long'] = ''

    raw = raw[
        raw['Award Number'].ne('')
        & raw['Task Number'].ne('')
        & raw['Purchase Order Number'].ne('')
        & raw['Expenditure Item Date'].notna()
    ].copy()

    raw['is_indirect'] = raw.apply(is_indirect_row, axis=1)
    raw['is_equipment'] = raw.apply(is_equipment_row, axis=1)
    raw['is_display_base'] = ~raw['is_indirect']
    raw['is_indirect_eligible_base'] = raw['is_display_base'] & (~raw['is_equipment'])

    log(f"[STEP] Parsed {len(raw)} usable commitment rows from raw data")
    return raw


def format_date_for_sheet(value):
    if pd.isna(value):
        return ''
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        return f"{value.month:02d}/{value.day:02d}/{value.year % 100:02d}"
    return str(value)


def build_entries_per_report_row(report_df, award, task):
    subset = report_df[(report_df['Award Number'] == award) & (report_df['Task Number'] == task)].copy()
    if subset.empty:
        return [], 0.0, 0.0, [f'{award}.{task}: no rows found in raw commitment report; will write zero']

    subset = subset.sort_values(
        ['Expenditure Item Date', 'Purchase Order Number', 'is_indirect', 'CODE', 'Expenditure Commitment Amount']
    ).reset_index(drop=True)

    report_total = round(float(subset['Expenditure Commitment Amount'].sum()), 2)
    base_rows = subset[subset['is_display_base']].copy().reset_index(drop=True)
    indirect_rows = subset[subset['is_indirect']].copy().reset_index(drop=True)

    entries = []
    for _, row in base_rows.iterrows():
        amount = round(safe_float(row['Expenditure Commitment Amount']), 2)
        entries.append({
            'date_display': format_date_for_sheet(row['Expenditure Item Date']),
            'date_key': pd.Timestamp(row['Expenditure Item Date']).normalize(),
            'code': str(row['CODE']),
            'po': str(row['Purchase Order Number']).strip(),
            'amount': amount,
            'indirect_amount': 0.0,
            'amount_plus_indirect': amount,
            'eligible_for_indirect': bool(row['is_indirect_eligible_base']),
        })

    warnings = []
    for _, ind in indirect_rows.iterrows():
        po = str(ind['Purchase Order Number']).strip()
        date_key = pd.Timestamp(ind['Expenditure Item Date']).normalize()
        indirect_amt = round(safe_float(ind['Expenditure Commitment Amount']), 2)

        exact_matches = [
            i for i, e in enumerate(entries)
            if e['eligible_for_indirect'] and e['po'] == po and e['date_key'] == date_key
        ]

        if len(exact_matches) == 1:
            idx = exact_matches[0]
            entries[idx]['indirect_amount'] = round(entries[idx]['indirect_amount'] + indirect_amt, 2)
            entries[idx]['amount_plus_indirect'] = round(entries[idx]['amount'] + entries[idx]['indirect_amount'], 2)
        elif len(exact_matches) > 1:
            # One indirect/F&A row can match multiple base commitment rows with the same PO and date.
            # The commitment sheet has no separate row for 56910 burden, so allocate the indirect
            # amount proportionally across the eligible base rows instead of dropping it.
            base_total = round(sum(entries[i]['amount'] for i in exact_matches), 2)
            if abs(base_total) <= TOLERANCE:
                warnings.append(
                    f"{award}.{task} indirect {indirect_amt:,.2f} on PO {po} date {format_date_for_sheet(ind['Expenditure Item Date'])} matched multiple rows but base total was zero; not auto-assigned"
                )
            else:
                remaining = indirect_amt
                for pos, idx in enumerate(exact_matches):
                    if pos == len(exact_matches) - 1:
                        share = round(remaining, 2)
                    else:
                        share = round(indirect_amt * entries[idx]['amount'] / base_total, 2)
                        remaining = round(remaining - share, 2)
                    entries[idx]['indirect_amount'] = round(entries[idx]['indirect_amount'] + share, 2)
                    entries[idx]['amount_plus_indirect'] = round(entries[idx]['amount'] + entries[idx]['indirect_amount'], 2)
                warnings.append(
                    f"{award}.{task} indirect {indirect_amt:,.2f} on PO {po} date {format_date_for_sheet(ind['Expenditure Item Date'])} was split proportionally across {len(exact_matches)} matching base rows"
                )
        else:
            warnings.append(
                f"{award}.{task} indirect {indirect_amt:,.2f} on PO {po} date {format_date_for_sheet(ind['Expenditure Item Date'])} could not be matched to exactly one eligible base row"
            )

    displayed_total = round(sum(e['amount_plus_indirect'] for e in entries), 2)
    difference = round(report_total - displayed_total, 2)
    if abs(difference) > TOLERANCE:
        warnings.append(
            f"{award}.{task} displayed total {displayed_total:,.2f} does not equal raw report total {report_total:,.2f}; difference = {difference:,.2f}"
        )

    return entries, report_total, displayed_total, warnings


def find_sheet_name(sheet_names, preferred):
    upper_map = {name.upper(): name for name in sheet_names}
    if preferred.upper() in upper_map:
        return upper_map[preferred.upper()]
    for name in sheet_names:
        if preferred.upper() in name.upper():
            return name
    return None


def locate_month_and_commit_row_openpyxl(ws, month_abbr_year):
    target_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        label = ws.cell(r, 1).value
        if isinstance(label, str) and 'CURRENT MO. COMMIT' in label.upper():
            target_row = r
            break
    if target_row is None:
        raise KeyError('Could not find row label containing CURRENT MO. COMMIT')

    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            if isinstance(value, datetime):
                if value.strftime('%b-%y').lower() == month_abbr_year:
                    return target_row, c, ws.cell(r, c).coordinate
            elif value is not None:
                text = str(value).strip().lower()
                if text == month_abbr_year or text == month_abbr_year.replace('-', '/'):
                    return target_row, c, ws.cell(r, c).coordinate
    raise KeyError(f'Could not find month column {month_abbr_year} on sheet {ws.title}')



def xl_col(n):
    result = ''
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def find_month_and_commit_row_excel(sheet, month_label):
    used = sheet.UsedRange
    max_row = min(used.Rows.Count, 60)
    max_col = used.Columns.Count
    target_row = None
    for r in range(1, max_row + 1):
        v = sheet.Cells(r, 1).Value
        if isinstance(v, str) and 'CURRENT MO. COMMIT' in v.upper():
            target_row = r
            break
    if target_row is None:
        raise KeyError('Could not find row label containing CURRENT MO. COMMIT')

    for r in range(1, min(30, used.Rows.Count) + 1):
        for c in range(1, max_col + 1):
            v = sheet.Cells(r, c).Value
            if v is None:
                continue
            if hasattr(v, 'strftime'):
                if v.strftime('%b-%y').lower() == month_label:
                    return target_row, c, f'{xl_col(c)}{r}'
            text = str(v).strip().lower()
            if text == month_label or text == month_label.replace('-', '/'):
                return target_row, c, f'{xl_col(c)}{r}'
    raise KeyError(f'Could not find month column {month_label} on sheet {sheet.Name}')


def workbook_award_task_excel(wb):
    for ws in wb.Worksheets:
        if '2025-2026' in ws.Name.upper():
            award = normalize_award(ws.Range('B3').Value)
            task = normalize_task(ws.Range('B5').Value)
            if award and task:
                return award, task, 'workbook 2025-2026 sheet'
    for ws in wb.Worksheets:
        if 'COMMITMENT' in ws.Name.upper():
            award = normalize_award(ws.Range('E3').Value)
            task = normalize_task(ws.Range('C5').Value)
            if not task and isinstance(ws.Range('C5').Value, str):
                _, task = parse_award_task_from_name(str(ws.Range('C5').Value))
            if award and task:
                return award, task, 'workbook COMMITMENT sheet'
    return None, None, None

def workbook_award_task_openpyxl(wb):
    sheet_name = find_sheet_name(wb.sheetnames, '2025-2026')
    if sheet_name:
        ws = wb[sheet_name]
        award = normalize_award(ws['B3'].value)
        task = normalize_task(ws['B5'].value)
        if award and task:
            return award, task, 'workbook 2025-2026 sheet'
    sheet_name = find_sheet_name(wb.sheetnames, 'COMMITMENT')
    if sheet_name:
        ws = wb[sheet_name]
        award = normalize_award(ws['E3'].value)
        task = normalize_task(ws['C5'].value)
        if not task and isinstance(ws['C5'].value, str):
            _, task = parse_award_task_from_name(str(ws['C5'].value))
        if award and task:
            return award, task, 'workbook COMMITMENT sheet'
    return None, None, None


def total_formula_text(left_value_col='E', right_value_col='K', start_row=6, end_row=35):
    return f'=SUM({left_value_col}{start_row}:{left_value_col}{end_row},{right_value_col}{start_row}:{right_value_col}{end_row})'

def total_formula_for_used_cells(used_value_cells, fallback_formula=None):
    """Build a formula that sums only the cells written for one task block."""
    if used_value_cells:
        return "=SUM(" + ",".join(used_value_cells) + ")"
    return fallback_formula or "=0"


def clear_block(ws, start_row):
    for row in range(start_row, start_row + 30):
        for col in list(range(1, 6)) + list(range(7, 12)):
            ws.cell(row, col).value = None
    for row in [start_row - 1, start_row + 30]:
        for col in list(range(1, 12)):
            ws.cell(row, col).value = None


def write_task_block(ws, entries, task, block_index=0):
    """Write one or two-column commitment block. block_index 0 begins at row 5/6."""
    header_row = 5 + block_index * 35
    data_start = header_row + 1
    data_end = data_start + 29
    total_row = data_end + 1

    clear_block(ws, data_start)

    ws[f'C{header_row}'] = f'TASK - {task}'

    used_value_cells = []
    for idx, entry in enumerate(entries):
        if idx < 30:
            row = data_start + idx
            cols = LEFT_BLOCK_COLS
        else:
            row = data_start + (idx - 30)
            cols = RIGHT_BLOCK_COLS
        ws[f'{cols[0]}{row}'] = entry['date_display']
        ws[f'{cols[1]}{row}'] = entry['code']
        ws[f'{cols[2]}{row}'] = str(entry['po'])
        ws[f'{cols[3]}{row}'] = float(entry['amount'])
        ws[f'{cols[4]}{row}'] = float(entry['amount_plus_indirect'])
        used_value_cells.append(f'{cols[4]}{row}')

    task_formula = total_formula_for_used_cells(used_value_cells, fallback_formula=total_formula_text('E', 'K', data_start, data_end))
    ws[f'E{header_row}'] = task_formula

    calc_total = round(sum(e['amount_plus_indirect'] for e in entries), 2)
    return calc_total, used_value_cells


def locate_graph_commit_cell_openpyxl(wb, month_label):
    graph_sheet_name = find_sheet_name(wb.sheetnames, '2025-2026')
    if not graph_sheet_name:
        raise KeyError('Could not find 2025-2026 sheet in workbook.')
    graph_ws = wb[graph_sheet_name]
    row_idx, col_idx, header_cell = locate_month_and_commit_row_openpyxl(graph_ws, month_label)
    target_cell = graph_ws.cell(row_idx, col_idx).coordinate
    return graph_ws, row_idx, col_idx, graph_ws.title, header_cell, target_cell


def fill_graph_commit_cell_openpyxl(wb, month_label, amount):
    graph_ws, row_idx, col_idx, graph_title, header_cell, target_cell = locate_graph_commit_cell_openpyxl(wb, month_label)
    graph_ws.cell(row_idx, col_idx).value = float(amount)
    return graph_title, header_cell, target_cell


def process_regular_openpyxl(path, hint_award, hint_task, report_df, month_label, sum_mode_awards):
    ext = Path(path).suffix.lower()
    keep_vba = ext == '.xlsm'
    log(f'[STEP] Opening workbook ({ext})')
    wb = load_workbook(path, data_only=False, keep_vba=keep_vba)

    award = task = source = None
    if hint_award and hint_task:
        award, task, source = hint_award, hint_task, 'shortcut/filename'
    if not (award and task):
        fa, ft = parse_award_task_from_name(Path(path).name)
        if fa and ft:
            award, task, source = fa, ft, 'filename'
    if not (award and task):
        award, task, source = workbook_award_task_openpyxl(wb)
    if not (award and task):
        raise KeyError('Could not identify award/task from filename or workbook cells.')
    log(f'[STEP] Award/task identified as {award}/{task} (from {source})')

    entries, report_total, displayed_total, warnings = build_entries_per_report_row(report_df, award, task)
    log(f'[STEP] Found {len(entries)} display row(s) for {award}/{task}; raw report total = {report_total:,.2f}')

    # Locate first; only fill after validation passes.
    graph_ws, graph_row_idx, graph_col_idx, graph_title, header_cell, target_cell = locate_graph_commit_cell_openpyxl(wb, month_label)

    wrote_detail = False
    calc_total = displayed_total
    used_value_cells = []
    # If a SUM workbook exists for this award, individual files only receive the graph total.
    if award not in sum_mode_awards:
        commit_sheet_name = find_sheet_name(wb.sheetnames, 'COMMITMENT')
        if not commit_sheet_name:
            raise KeyError('Could not find COMMITMENT sheet in workbook.')
        if len(entries) > 60:
            raise RuntimeError(f'{award}.{task} has {len(entries)} commitment lines, but one task block holds only 60.')
        commit_ws = wb[commit_sheet_name]
        calc_total, used_value_cells = write_task_block(commit_ws, entries, task, block_index=0)
        wrote_detail = True
    else:
        warnings.append(f'{award}.{task}: SUM file exists for award {award}; only filled CURRENT MO. COMMIT cell in this task workbook.')

    diff = round(report_total - calc_total, 2)
    status = 'OK' if abs(diff) <= TOLERANCE else 'CHECK'
    if abs(diff) > TOLERANCE:
        warnings.append(f'{award}.{task} validation: displayed total {calc_total:,.2f} vs raw report total {report_total:,.2f}')
        graph_ws.cell(graph_row_idx, graph_col_idx).value = None
        warnings.append(f'{award}.{task}: CURRENT MO. COMMIT cell {target_cell} was not filled because validation failed.')
    else:
        graph_ws.cell(graph_row_idx, graph_col_idx).value = float(report_total)

    wb.save(path)

    return {
        'path': path, 'award': award, 'task': task, 'file_type': 'TASK',
        'entries': len(entries), 'total': report_total, 'calc_total': calc_total,
        'difference': diff, 'status': status, 'target_cell': target_cell,
        'used_value_cells': ','.join(used_value_cells), 'warnings': warnings,
        'detail_written': wrote_detail,
    }



def process_sum_openpyxl(path, report_df, month_label, selected_tasks=None):
    award = parse_award_sum_from_name(Path(path).name)
    if not award:
        raise KeyError('Could not identify SUM award from filename.')

    ext = Path(path).suffix.lower()
    keep_vba = ext == '.xlsm'
    log(f'[STEP] Opening SUM workbook ({ext}) for award {award}')
    wb = load_workbook(path, data_only=False, keep_vba=keep_vba)

    commit_sheet_name = find_sheet_name(wb.sheetnames, 'COMMITMENT')
    if not commit_sheet_name:
        raise KeyError('Could not find COMMITMENT sheet in SUM workbook.')
    commit_ws = wb[commit_sheet_name]

    raw_tasks = set(report_df.loc[report_df['Award Number'] == award, 'Task Number'].dropna().astype(str).unique())
    if selected_tasks is None:
        tasks = raw_tasks
        skipped_tasks = set()
        selected_missing_from_raw = set()
    else:
        selected_tasks = set(str(t) for t in selected_tasks)
        tasks = raw_tasks & selected_tasks
        skipped_tasks = raw_tasks - selected_tasks
        selected_missing_from_raw = selected_tasks - raw_tasks

    tasks = sorted(tasks, key=lambda x: int(x) if str(x).isdigit() else str(x))
    if not tasks:
        raise KeyError(
            f'No selected task workbooks in the input folder matched raw commitment rows for SUM award {award}. '
            f'Selected tasks: {sorted(selected_tasks) if selected_tasks is not None else "ALL"}; '
            f'raw tasks: {sorted(raw_tasks)}.'
        )

    total_report = 0.0
    total_displayed = 0.0
    warnings = []
    used_cells_all = []
    task_summaries = []
    entries_count = 0

    if selected_tasks is not None:
        warnings.append(f'{award}.SUM: only tasks present in the selected input folder were written: {", ".join(tasks)}.')
        if skipped_tasks:
            warnings.append(f'{award}.SUM: skipped raw-report task(s) not present in the selected folder: {", ".join(sorted(skipped_tasks, key=lambda x: int(x) if str(x).isdigit() else str(x)))}.')
        if selected_missing_from_raw:
            warnings.append(f'{award}.SUM: selected folder had task file(s) with no raw commitment rows: {", ".join(sorted(selected_missing_from_raw, key=lambda x: int(x) if str(x).isdigit() else str(x)))}.')

    # Clear likely old task sections. This handles several stacked blocks.
    for block_index in range(max(8, len(tasks))):
        clear_block(commit_ws, 6 + block_index * 35)

    for i, task in enumerate(tasks):
        entries, report_total, displayed_total, task_warnings = build_entries_per_report_row(report_df, award, task)
        entries_count += len(entries)
        if len(entries) > 60:
            warnings.append(f'{award}.{task} has {len(entries)} commitment lines; only first 60 can fit in one block.')
            entries = entries[:60]
        calc_total, used_cells = write_task_block(commit_ws, entries, task, block_index=i)
        total_report = round(total_report + report_total, 2)
        total_displayed = round(total_displayed + calc_total, 2)
        warnings.extend(task_warnings)
        used_cells_all.extend(used_cells)
        task_summaries.append(f'{task}:{report_total:.2f}')

    # Locate SUM workbook graph cell, but only fill it after validation passes.
    graph_ws, graph_row_idx, graph_col_idx, graph_title, header_cell, target_cell = locate_graph_commit_cell_openpyxl(wb, month_label)

    diff = round(total_report - total_displayed, 2)
    status = 'OK' if abs(diff) <= TOLERANCE else 'CHECK'
    if abs(diff) > TOLERANCE:
        warnings.append(f'{award}.SUM validation: displayed total {total_displayed:,.2f} vs raw report total {total_report:,.2f}')
        graph_ws.cell(graph_row_idx, graph_col_idx).value = None
        warnings.append(f'{award}.SUM: CURRENT MO. COMMIT cell {target_cell} was not filled because validation failed.')
    else:
        graph_ws.cell(graph_row_idx, graph_col_idx).value = float(total_report)

    wb.save(path)

    return {
        'path': path, 'award': award, 'task': 'SUM', 'file_type': 'SUM',
        'entries': entries_count,
        'total': total_report, 'calc_total': total_displayed,
        'difference': diff, 'status': status, 'target_cell': target_cell,
        'used_value_cells': ','.join(used_cells_all), 'warnings': warnings,
        'detail_written': True, 'task_totals': '; '.join(task_summaries),
        'tasks_found': len(tasks),
    }



def clear_block_excel(ws, start_row):
    end_row = start_row + 29
    ws.Range(f'A{start_row}:E{end_row}').ClearContents()
    ws.Range(f'G{start_row}:K{end_row}').ClearContents()
    for row in (start_row - 1, start_row + 30):
        ws.Range(f'A{row}:K{row}').ClearContents()


def write_task_block_excel(ws, entries, task, block_index=0):
    header_row = 5 + block_index * 35
    data_start = header_row + 1
    data_end = data_start + 29
    total_row = data_end + 1

    clear_block_excel(ws, data_start)
    ws.Range(f'C{header_row}').Value = f'TASK - {task}'

    used_value_cells = []
    for idx, entry in enumerate(entries):
        if idx < 30:
            row = data_start + idx
            cols = LEFT_BLOCK_COLS
        else:
            row = data_start + (idx - 30)
            cols = RIGHT_BLOCK_COLS
        ws.Range(f'{cols[0]}{row}').Value = entry['date_display']
        ws.Range(f'{cols[1]}{row}').Value = entry['code']
        ws.Range(f'{cols[2]}{row}').Value = str(entry['po'])
        ws.Range(f'{cols[3]}{row}').Value = float(entry['amount'])
        ws.Range(f'{cols[4]}{row}').Value = float(entry['amount_plus_indirect'])
        used_value_cells.append(f'{cols[4]}{row}')

    task_formula = total_formula_for_used_cells(used_value_cells, fallback_formula=total_formula_text('E', 'K', data_start, data_end))
    ws.Range(f'E{header_row}').Formula = task_formula

    calc_total = round(sum(e['amount_plus_indirect'] for e in entries), 2)
    return calc_total, used_value_cells


def locate_graph_commit_cell_excel(wb, month_label):
    graph_ws = None
    for ws in wb.Worksheets:
        if '2025-2026' in ws.Name.upper():
            graph_ws = ws
            break
    if graph_ws is None:
        raise KeyError('Could not find 2025-2026 sheet in workbook.')
    row_idx, col_idx, header_cell = find_month_and_commit_row_excel(graph_ws, month_label)
    target_cell = f'{xl_col(col_idx)}{row_idx}'
    return graph_ws, target_cell, graph_ws.Name, header_cell


def fill_graph_commit_cell_excel(wb, month_label, amount):
    graph_ws, target_cell, graph_title, header_cell = locate_graph_commit_cell_excel(wb, month_label)
    graph_ws.Range(target_cell).Value = float(amount)
    return graph_title, header_cell, target_cell


def process_regular_excel_com(path, hint_award, hint_task, report_df, month_label, sum_mode_awards, recalc=False):
    try:
        import win32com.client
    except Exception as exc:
        raise RuntimeError('pywin32 is required to edit .xls files and resolve Windows shortcuts. Run: python -m pip install pywin32') from exc

    log('[STEP] Opening legacy .xls workbook via Excel COM')
    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(os.path.abspath(path), UpdateLinks=False, ReadOnly=False)
        award = task = source = None
        if hint_award and hint_task:
            award, task, source = hint_award, hint_task, 'shortcut/filename'
        if not (award and task):
            fa, ft = parse_award_task_from_name(Path(path).name)
            if fa and ft:
                award, task, source = fa, ft, 'filename'
        if not (award and task):
            award, task, source = workbook_award_task_excel(wb)
        if not (award and task):
            raise KeyError('Could not identify award/task from shortcut name, filename, or workbook cells.')
        log(f'[STEP] Award/task identified as {award}/{task} (from {source})')

        entries, report_total, displayed_total, warnings = build_entries_per_report_row(report_df, award, task)
        log(f'[STEP] Found {len(entries)} display row(s) for {award}/{task}; raw report total = {report_total:,.2f}')

        graph_ws, target_cell, graph_title, header_cell = locate_graph_commit_cell_excel(wb, month_label)

        wrote_detail = False
        calc_total = displayed_total
        used_value_cells = []
        if award not in sum_mode_awards:
            commit_ws = None
            for ws in wb.Worksheets:
                if 'COMMITMENT' in ws.Name.upper():
                    commit_ws = ws
                    break
            if commit_ws is None:
                raise KeyError('Could not find COMMITMENT sheet in workbook.')
            if len(entries) > 60:
                raise RuntimeError(f'{award}.{task} has {len(entries)} commitment lines, but one task block holds only 60.')
            calc_total, used_value_cells = write_task_block_excel(commit_ws, entries, task, block_index=0)
            wrote_detail = True
        else:
            warnings.append(f'{award}.{task}: SUM file exists for award {award}; only filled CURRENT MO. COMMIT cell in this task workbook.')

        if recalc:
            excel.CalculateFull()
        diff = round(report_total - calc_total, 2)
        status = 'OK' if abs(diff) <= TOLERANCE else 'CHECK'
        if abs(diff) > TOLERANCE:
            warnings.append(f'{award}.{task} validation: displayed total {calc_total:,.2f} vs raw report total {report_total:,.2f}')
            graph_ws.Range(target_cell).ClearContents()
            warnings.append(f'{award}.{task}: CURRENT MO. COMMIT cell {target_cell} was not filled because validation failed.')
        else:
            graph_ws.Range(target_cell).Value = float(report_total)
        wb.Save()
        return {
            'path': path, 'award': award, 'task': task, 'file_type': 'TASK',
            'entries': len(entries), 'total': report_total, 'calc_total': calc_total,
            'difference': diff, 'status': status, 'target_cell': target_cell,
            'used_value_cells': ','.join(used_value_cells), 'warnings': warnings,
            'detail_written': wrote_detail,
        }
    finally:
        if wb is not None:
            wb.Close(SaveChanges=True)
        excel.Quit()


def process_sum_excel_com(path, report_df, month_label, selected_tasks=None, recalc=False):
    try:
        import win32com.client
    except Exception as exc:
        raise RuntimeError('pywin32 is required to edit .xls files and resolve Windows shortcuts. Run: python -m pip install pywin32') from exc

    award = parse_award_sum_from_name(Path(path).name)
    if not award:
        raise KeyError('Could not identify SUM award from filename.')

    log(f'[STEP] Opening legacy .xls SUM workbook via Excel COM for award {award}')
    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(os.path.abspath(path), UpdateLinks=False, ReadOnly=False)
        commit_ws = None
        for ws in wb.Worksheets:
            if 'COMMITMENT' in ws.Name.upper():
                commit_ws = ws
                break
        if commit_ws is None:
            raise KeyError('Could not find COMMITMENT sheet in SUM workbook.')

        raw_tasks = set(report_df.loc[report_df['Award Number'] == award, 'Task Number'].dropna().astype(str).unique())
        if selected_tasks is None:
            tasks = raw_tasks
            skipped_tasks = set()
            selected_missing_from_raw = set()
        else:
            selected_tasks = set(str(t) for t in selected_tasks)
            tasks = raw_tasks & selected_tasks
            skipped_tasks = raw_tasks - selected_tasks
            selected_missing_from_raw = selected_tasks - raw_tasks

        tasks = sorted(tasks, key=lambda x: int(x) if str(x).isdigit() else str(x))
        if not tasks:
            raise KeyError(
                f'No selected task workbooks in the input folder matched raw commitment rows for SUM award {award}. '
                f'Selected tasks: {sorted(selected_tasks) if selected_tasks is not None else "ALL"}; '
                f'raw tasks: {sorted(raw_tasks)}.'
            )

        total_report = 0.0
        total_displayed = 0.0
        warnings = []
        used_cells_all = []
        task_summaries = []
        entries_count = 0

        if selected_tasks is not None:
            warnings.append(f'{award}.SUM: only tasks present in the selected input folder were written: {", ".join(tasks)}.')
            if skipped_tasks:
                warnings.append(f'{award}.SUM: skipped raw-report task(s) not present in the selected folder: {", ".join(sorted(skipped_tasks, key=lambda x: int(x) if str(x).isdigit() else str(x)))}.')
            if selected_missing_from_raw:
                warnings.append(f'{award}.SUM: selected folder had task file(s) with no raw commitment rows: {", ".join(sorted(selected_missing_from_raw, key=lambda x: int(x) if str(x).isdigit() else str(x)))}.')

        for block_index in range(max(8, len(tasks))):
            clear_block_excel(commit_ws, 6 + block_index * 35)

        for i, task in enumerate(tasks):
            entries, report_total, displayed_total, task_warnings = build_entries_per_report_row(report_df, award, task)
            entries_count += len(entries)
            if len(entries) > 60:
                warnings.append(f'{award}.{task} has {len(entries)} commitment lines; only first 60 can fit in one block.')
                entries = entries[:60]
            calc_total, used_cells = write_task_block_excel(commit_ws, entries, task, block_index=i)
            total_report = round(total_report + report_total, 2)
            total_displayed = round(total_displayed + calc_total, 2)
            warnings.extend(task_warnings)
            used_cells_all.extend(used_cells)
            task_summaries.append(f'{task}:{report_total:.2f}')

        graph_ws, target_cell, graph_title, header_cell = locate_graph_commit_cell_excel(wb, month_label)
        if recalc:
            excel.CalculateFull()
        diff = round(total_report - total_displayed, 2)
        status = 'OK' if abs(diff) <= TOLERANCE else 'CHECK'
        if abs(diff) > TOLERANCE:
            warnings.append(f'{award}.SUM validation: displayed total {total_displayed:,.2f} vs raw report total {total_report:,.2f}')
            graph_ws.Range(target_cell).ClearContents()
            warnings.append(f'{award}.SUM: CURRENT MO. COMMIT cell {target_cell} was not filled because validation failed.')
        else:
            graph_ws.Range(target_cell).Value = float(total_report)
        wb.Save()
        return {
            'path': path, 'award': award, 'task': 'SUM', 'file_type': 'SUM',
            'entries': entries_count, 'total': total_report, 'calc_total': total_displayed,
            'difference': diff, 'status': status, 'target_cell': target_cell,
            'used_value_cells': ','.join(used_cells_all), 'warnings': warnings,
            'detail_written': True, 'task_totals': '; '.join(task_summaries),
            'tasks_found': len(tasks),
        }
    finally:
        if wb is not None:
            wb.Close(SaveChanges=True)
        excel.Quit()



def process_one(input_path, report_df, month_label, sum_mode_awards, selected_tasks_by_award):
    display_path = input_path
    target_path = input_path
    hint_award = hint_task = None
    if Path(input_path).suffix.lower() == '.lnk':
        hint_award, hint_task = parse_award_task_from_name(Path(input_path).name)
        target_path = resolve_shortcut(input_path)
        log(f'[STEP] Resolved shortcut: {input_path} -> {target_path}')

    ext = Path(target_path).suffix.lower()
    if ext not in EXCEL_EXTS:
        raise RuntimeError(f'Skipping unsupported target type: {target_path}')

    log(f'\\n[FILE] {target_path}')
    if ext == '.xls':
        if is_sum_file(target_path):
            award = parse_award_sum_from_name(Path(target_path).name)
            result = process_sum_excel_com(target_path, report_df, month_label, selected_tasks=selected_tasks_by_award.get(award), recalc=True)
        else:
            result = process_regular_excel_com(target_path, hint_award, hint_task, report_df, month_label, sum_mode_awards, recalc=True)
    elif is_sum_file(target_path):
        award = parse_award_sum_from_name(Path(target_path).name)
        result = process_sum_openpyxl(target_path, report_df, month_label, selected_tasks=selected_tasks_by_award.get(award))
    else:
        result = process_regular_openpyxl(target_path, hint_award, hint_task, report_df, month_label, sum_mode_awards)
    result['display_path'] = display_path
    return result


def detect_sum_mode_and_selected_tasks(files):
    """Detect every AWARD.SUM workbook and every AWARD.TASK workbook selected by the user.

    SUM commitment sheets should only include tasks that are represented by task
    workbooks/shortcuts in the selected input set. This prevents a SUM workbook
    from pulling every raw-report task for the award when the folder only contains
    a subset such as 601 and 603.
    """
    sum_mode_awards = set()
    selected_tasks_by_award = {}
    notes = []

    for path in files:
        candidates = [path]
        if Path(path).suffix.lower() == '.lnk':
            try:
                candidates.append(resolve_shortcut(path))
            except Exception as exc:
                notes.append(f'Could not inspect shortcut for SUM/task detection: {path} ({exc})')

        # First detect SUM files. SUM files are not task selections.
        detected_sum_award = None
        for candidate in candidates:
            award = parse_award_sum_from_name(Path(candidate).name)
            if award:
                detected_sum_award = award
                break
        if detected_sum_award:
            sum_mode_awards.add(detected_sum_award)
            notes.append(f'SUM detected: {detected_sum_award} from {path}')
            continue

        # Then detect task files selected in the folder.
        detected_award = detected_task = None
        for candidate in candidates:
            detected_award, detected_task = parse_award_task_from_name(Path(candidate).name)
            if detected_award and detected_task:
                break
        if detected_award and detected_task:
            selected_tasks_by_award.setdefault(detected_award, set()).add(str(detected_task))
            notes.append(f'Task selected: {detected_award}.{detected_task} from {path}')

    return sum_mode_awards, selected_tasks_by_award, notes



def choose_default_output_dir(inputs):
    for item in inputs:
        p = Path(strip_quotes(item))
        if p.is_dir():
            return str(p)
        if p.is_file():
            return str(p.parent)
    return os.getcwd()


def write_processing_reports(output_dir, rows):
    ensure_dir(output_dir)
    xlsx_path = os.path.join(output_dir, 'commit_fill_report.xlsx')

    headers = [
        'status', 'file_type', 'path', 'award', 'task', 'tasks_found', 'entries',
        'report_total', 'displayed_total', 'difference', 'target_cell',
        'detail_written', 'task_totals', 'used_value_cells', 'notes'
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Processing Report'
    ws.append(headers)

    for row in rows:
        ws.append([row.get(h, '') for h in headers])

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(bottom=thin)

    status_col = headers.index('status') + 1
    for r in range(2, ws.max_row + 1):
        status = str(ws.cell(r, status_col).value or '').upper()
        fill = None
        if status == 'OK':
            fill = PatternFill('solid', fgColor='E2F0D9')
        elif status == 'CHECK':
            fill = PatternFill('solid', fgColor='FFF2CC')
        elif status == 'FAIL':
            fill = PatternFill('solid', fgColor='FCE4D6')
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill

    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            max_len = max(max_len, len(str(ws.cell(row_idx, col_idx).value or '')))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)

    log(f'Excel processing report written to: {xlsx_path}')



def read_commitment_report(report_path):
    """Read the raw commitment report.

    If the workbook has both a filtered drill-down sheet (for example Detail1)
    and the full raw export (for example Sheet1), choose the matching sheet/header
    with the largest number of usable raw rows.
    """
    log(f"[STEP] Reading raw commitment report: {report_path}")

    xl = pd.ExcelFile(report_path)
    best = None
    best_guess = None

    for sheet_name in xl.sheet_names:
        for header_row in range(0, 31):
            try:
                candidate = pd.read_excel(report_path, sheet_name=sheet_name, header=header_row)
            except Exception:
                continue
            candidate.columns = make_unique_columns(candidate.columns)
            mapping, missing = find_report_columns(candidate.columns)
            found_count = 6 - len(missing)
            if best_guess is None or found_count > best_guess['found_count']:
                best_guess = {
                    'sheet': sheet_name,
                    'header_row': header_row,
                    'found_count': found_count,
                    'missing': missing,
                    'columns': [str(c) for c in candidate.columns[:60]],
                }
            if missing:
                continue

            test = candidate.rename(columns={actual: standard for standard, actual in mapping.items()}).copy()
            first_col = test.columns[0]
            test = test[~test[first_col].astype(str).str.strip().isin(['Grand Total', 'Rows 1 - 90 (All Rows)'])].copy()

            for col in ['Project Number', 'Task Number', 'Award Number',
                        'Expenditure High Level Group Description', 'Expenditure Type Long Name',
                        'Source System Description']:
                if col in test.columns:
                    test[col] = test[col].ffill()

            test['Award Number'] = test['Award Number'].map(normalize_award)
            test['Task Number'] = test['Task Number'].map(normalize_task)
            test['Purchase Order Number'] = test['Purchase Order Number'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
            test['Expenditure Item Date'] = test['Expenditure Item Date'].map(excel_serial_to_datetime)
            test['Expenditure Commitment Amount'] = pd.to_numeric(test['Expenditure Commitment Amount'], errors='coerce').fillna(0.0)

            usable_mask = (
                test['Award Number'].ne('')
                & test['Task Number'].ne('')
                & test['Purchase Order Number'].ne('')
                & test['Expenditure Item Date'].notna()
            )
            usable_count = int(usable_mask.sum())

            if best is None or usable_count > best['usable_count']:
                best = {
                    'raw': candidate,
                    'mapping': mapping,
                    'sheet': sheet_name,
                    'header_row': header_row,
                    'usable_count': usable_count,
                }

    if best is None:
        if best_guess:
            raise KeyError(
                "Raw commitment report columns were not found automatically. "
                f"Best guess: sheet '{best_guess['sheet']}', header row {best_guess['header_row'] + 1}, "
                f"found {best_guess['found_count']}/6 columns. Missing: {', '.join(best_guess['missing'])}. "
                f"Columns seen: {best_guess['columns']}"
            )
        raise KeyError(f"Raw commitment report columns were not found automatically. Tried sheets {xl.sheet_names}.")

    raw = best['raw']
    used_mapping = best['mapping']
    log(f"[STEP] Using sheet '{best['sheet']}' with header row {best['header_row'] + 1}")

    rename_map = {actual: standard for standard, actual in used_mapping.items()}
    raw = raw.rename(columns=rename_map).copy()

    first_col = raw.columns[0]
    raw = raw[~raw[first_col].astype(str).str.strip().isin(['Grand Total', 'Rows 1 - 90 (All Rows)'])].copy()

    ffill_cols = [
        'Project Number', 'Task Number', 'Award Number',
        'Expenditure High Level Group Description', 'Expenditure Type Long Name',
        'Source System Description'
    ]
    for col in ffill_cols:
        if col in raw.columns:
            raw[col] = raw[col].ffill()

    raw['Award Number'] = raw['Award Number'].map(normalize_award)
    raw['Task Number'] = raw['Task Number'].map(normalize_task)
    raw['Purchase Order Number'] = raw['Purchase Order Number'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    raw['Expenditure Item Date'] = raw['Expenditure Item Date'].map(excel_serial_to_datetime)
    raw['Expenditure Commitment Amount'] = pd.to_numeric(raw['Expenditure Commitment Amount'], errors='coerce').fillna(0.0)
    raw['CODE'] = raw['Expenditure Type Long Name'].map(extract_code)

    if 'Expenditure High Level Group Description' in raw.columns:
        raw['High Level'] = raw['Expenditure High Level Group Description'].fillna('').astype(str)
    else:
        raw['High Level'] = ''

    if 'Expenditure Type Long Name' in raw.columns:
        raw['Type Long'] = raw['Expenditure Type Long Name'].fillna('').astype(str)
    else:
        raw['Type Long'] = ''

    raw = raw[
        raw['Award Number'].ne('')
        & raw['Task Number'].ne('')
        & raw['Purchase Order Number'].ne('')
        & raw['Expenditure Item Date'].notna()
    ].copy()

    raw['is_indirect'] = raw.apply(is_indirect_row, axis=1)
    raw['is_equipment'] = raw.apply(is_equipment_row, axis=1)
    raw['is_display_base'] = ~raw['is_indirect']
    raw['is_indirect_eligible_base'] = raw['is_display_base'] & (~raw['is_equipment'])

    log(f"[STEP] Parsed {len(raw)} usable commitment rows from raw data")
    return raw


def main():
    parser = argparse.ArgumentParser(
    description="Fill commitment detail tables and monthly commitment cells from a source commitment report."
    )
    parser.add_argument('--commit-report', help='Path to the raw commitment report workbook (.xlsx).')
    parser.add_argument('--inputs', nargs='+', help='One or more input Excel files, shortcuts, or directories.')
    parser.add_argument('--month', help='Month label like mar-2026. Used to locate the graph column to fill.')
    parser.add_argument('--output-dir', default=None, help='Where to write commit_fill_report.xlsx.')
    parser.add_argument('--include-subfolders', action='store_true', help='Recurse into input directories.')
    parser.add_argument('--no-report', action='store_true', help='Do not create the Excel/CSV summary files.')
    args = parser.parse_args()

    if not args.commit_report:
        args.commit_report = strip_quotes(input('Enter FULL path to source commitment report Excel: '))
    else:
        args.commit_report = strip_quotes(args.commit_report)

    if not args.inputs:
        user_input = strip_quotes(input('Enter FULL path to Input folder/files: '))
        args.inputs = [user_input]
    else:
        args.inputs = [strip_quotes(x) for x in args.inputs]

    if not args.month:
        args.month = strip_quotes(input('Enter month (e.g. mar-2026): '))

    if args.output_dir:
        args.output_dir = strip_quotes(args.output_dir)
    else:
        args.output_dir = choose_default_output_dir(args.inputs)

    print('\n=== RUNNING WITH ===')
    print(f'Report: {args.commit_report}')
    print(f'Inputs: {args.inputs}')
    print(f'Month:  {args.month}')
    print(f'Output report folder: {args.output_dir}')
    print('=' * 30)

    try:
        _, _, month_label, _ = parse_month(args.month)
    except Exception as exc:
        print(f'[FAIL] Month could not be parsed: {exc}')
        sys.exit(1)

    files = iter_input_files(args.inputs, include_subfolders=args.include_subfolders)
    log(f'[STEP] Found {len(files)} workbook/shortcut file(s) to process')
    if not files:
        print('No Excel files were found in the provided inputs.')
        sys.exit(1)

    report_rows = []
    try:
        report_df = read_commitment_report(args.commit_report)
    except Exception as exc:
        reason = f'Raw commitment report could not be read, so no workbooks were updated. Reason: {exc}'
        print(f'[FAIL] {reason}')
        for path in files:
            award, task = parse_award_task_from_name(Path(path).name)
            if is_sum_file(path):
                award = parse_award_sum_from_name(Path(path).name)
                task = 'SUM'
            report_rows.append({
                'status': 'FAIL', 'file_type': 'SUM' if is_sum_file(path) else 'TASK',
                'path': path, 'award': award or '', 'task': task or '',
                'notes': reason
            })
        if not args.no_report:
            write_processing_reports(args.output_dir, report_rows)
        sys.exit(1)

    # Detect SUM workbooks for EVERY award in the selected folder.
    # Also detect the task workbooks selected in the folder so SUM files only
    # write task blocks for the tasks the user actually included.
    sum_mode_awards, selected_tasks_by_award, sum_detection_notes = detect_sum_mode_and_selected_tasks(files)

    if sum_mode_awards:
        log(f"[STEP] SUM workbook mode detected for award(s): {', '.join(sorted(sum_mode_awards))}")
        for award in sorted(sum_mode_awards):
            selected = sorted(selected_tasks_by_award.get(award, set()), key=lambda x: int(x) if str(x).isdigit() else str(x))
            if selected:
                log(f"[STEP] {award}.SUM will include only selected folder task(s): {', '.join(selected)}")
            else:
                log(f"[WARN] {award}.SUM was detected, but no {award}.TASK workbook was detected in the selected folder.")
        for note in sum_detection_notes:
            log(f"[STEP] {note}")
    else:
        log("[STEP] No SUM workbooks detected. Each task workbook will receive its own COMMITMENT detail table.")

    failures = []
    needs_review = []

    for path in files:
        try:
            result = process_one(path, report_df, month_label, sum_mode_awards, selected_tasks_by_award)
            notes = '; '.join(result.get('warnings', []))
            report_rows.append({
                'status': result['status'],
                'file_type': result.get('file_type', ''),
                'path': result['path'],
                'award': result['award'],
                'task': result['task'],
                'tasks_found': result.get('tasks_found', ''),
                'entries': result.get('entries', ''),
                'report_total': f"{result['total']:.2f}",
                'displayed_total': f"{result['calc_total']:.2f}",
                'difference': f"{result['difference']:.2f}",
                'target_cell': result['target_cell'],
                'detail_written': 'YES' if result.get('detail_written') else 'NO',
                'task_totals': result.get('task_totals', ''),
                'used_value_cells': result.get('used_value_cells', ''),
                'notes': notes,
            })
            if result['status'] != 'OK':
                needs_review.append(f"{result['award']}.{result['task']}")
            print(
                f"[{result['status']}] {Path(result['path']).name} | "
                f"{result['award']}/{result['task']} | filled {result['target_cell']} with {result['total']:,.2f}"
            )
        except Exception as exc:
            award, task = parse_award_task_from_name(Path(path).name)
            file_type = 'TASK'
            if is_sum_file(path):
                award = parse_award_sum_from_name(Path(path).name)
                task = 'SUM'
                file_type = 'SUM'
            failures.append(f'{path}: {exc}')
            if award and task:
                needs_review.append(f'{award}.{task}')
            report_rows.append({
                'status': 'FAIL',
                'file_type': file_type,
                'path': path,
                'award': award or '',
                'task': task or '',
                'notes': str(exc),
            })
            print(f'[FAIL] {path}: {exc}')

    if not args.no_report:
        write_processing_reports(args.output_dir, report_rows)

    print('\nAward.task items needing review:')
    if needs_review:
        for item in sorted(set(needs_review)):
            print(item)
    else:
        print('None')

    if failures:
        print('\nFailures:')
        for item in failures:
            print(f'  - {item}')
        sys.exit(1)


if __name__ == '__main__':
    main()
