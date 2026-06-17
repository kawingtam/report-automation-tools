#!/usr/bin/env python3
"""
Fill CURRENT MO. EXPENSE / balance graph workbooks from an Expenditure Summary workbook.

Interactive version:
- Prompts for raw Expenditure Summary workbook, input folder/files, and month if arguments are not supplied.
- Edits workbooks in place only after validation passes.
- Supports .xls/.xlsx/.xlsm and .lnk shortcuts on Windows through Microsoft Excel COM.
- Handles AWARD.SUM workbooks by summing only task files present in the selected folder.
- Writes one pretty Excel processing report: balance_fill_report.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TOLERANCE = 0.01
EXCEL_EXTS = {".xls", ".xlsx", ".xlsm"}
INPUT_EXTS = EXCEL_EXTS | {".lnk"}


@dataclass
class BalanceTaskData:
    award: str
    task: str
    month_actual: float
    ptd_actual: float
    budget: float
    remaining_balance: float


@dataclass
class WorkItem:
    input_path: Path
    workbook_path: Path


def log(msg: str) -> None:
    print(msg)


def strip_quotes(value) -> str:
    return str(value or "").strip().strip('"').strip("'")


def normalize_text(value) -> str:
    return str(value or "").strip()


def normalize_award(value) -> str:
    s = normalize_text(value).upper()
    m = re.search(r"([A-Z0-9]{5})", s)
    return m.group(1) if m else s


def normalize_task(value) -> str:
    s = normalize_text(value)
    if s.endswith(".0"):
        s = s[:-2]
    m = re.search(r"(\d{1,5})", s)
    return m.group(1) if m else s


def safe_float(value) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def parse_month_input(month_text: str) -> Tuple[int, int, str, str]:
    cleaned = strip_quotes(month_text).replace("_", "-").replace("/", "-")
    cleaned = re.sub(r"\s+", " ", cleaned)
    formats = [
        "%b-%Y", "%b-%y", "%B-%Y", "%B-%y",
        "%b %Y", "%b %y", "%B %Y", "%B %y",
        "%Y-%m", "%Y %m", "%m-%Y", "%m-%y",
    ]
    for fmt in formats:
        try:
            parsed = dt.datetime.strptime(cleaned, fmt)
            return parsed.year, parsed.month, parsed.strftime("%b-%Y").upper(), parsed.strftime("%b-%y").lower()
        except ValueError:
            pass
    raise ValueError(f"Could not parse month '{month_text}'. Try values like mar-2026 or 2026-03.")


def month_matches(value, year: int, month: int) -> bool:
    if isinstance(value, (dt.datetime, dt.date)):
        return value.year == year and value.month == month
    text = normalize_text(value).lower()
    if not text:
        return False
    for fmt in ("%b-%Y", "%b-%y", "%B-%Y", "%B-%y", "%Y-%m", "%m-%Y", "%m-%y"):
        try:
            parsed = dt.datetime.strptime(text.replace("/", "-"), fmt)
            return parsed.year == year and parsed.month == month
        except ValueError:
            continue
    return False


def is_any_month_header(value) -> bool:
    """Return True if a cell looks like any month/year header, e.g. FEB-2026."""
    if isinstance(value, (dt.datetime, dt.date)):
        return True
    text = normalize_text(value).lower().replace("/", "-")
    if not text:
        return False
    for fmt in ("%b-%Y", "%b-%y", "%B-%Y", "%B-%y", "%Y-%m", "%m-%Y", "%m-%y"):
        try:
            dt.datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def clean_header(value) -> str:
    return re.sub(r"\s+", " ", normalize_text(value)).strip()


def read_balance_report(report_path: str, month_header_key: str) -> Dict[Tuple[str, str], BalanceTaskData]:
    """Read either:
    1) a pivot-style Expenditure Summary with Row Labels award/task hierarchy, or
    2) a raw/detail Expenditure Summary with Award Number and Task Number columns.

    Some Oracle/Excel exports have incomplete worksheet dimensions in read-only mode.
    This reader scans rows directly instead of relying on ws.max_row/ws.max_column.
    """
    log(f"[STEP] Reading balance source workbook: {report_path}")
    wb = load_workbook(report_path, data_only=True, read_only=True)

    best = None
    for ws in wb.worksheets:
        preview_rows = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            preview_rows.append((row_idx, list(row)))
            if row_idx >= 40:
                break

        for header_row, row_values in preview_rows:
            headers = [clean_header(v) for v in row_values]
            upper = [h.upper() for h in headers]

            row_label_col = None
            award_col = None
            task_col = None
            month_actual_col = None
            budget_col = None
            ptd_col = None
            remaining_col = None

            for idx, h in enumerate(upper, start=1):
                if h in {"ROW LABELS", "ROW LABEL", "AWARD/TASK", "AWARD TASK"}:
                    row_label_col = idx
                if h == "AWARD NUMBER" and award_col is None:
                    award_col = idx
                if h == "TASK NUMBER" and task_col is None:
                    task_col = idx
                if month_header_key in h and "ACTUAL EXPENDITURE AMOUNT" in h:
                    month_actual_col = idx
                if "BUDGET AMOUNT" in h and budget_col is None:
                    budget_col = idx
                if ("FTD OR PTD ACTUAL EXPENDITURE AMOUNT" in h or "PTD ACTUAL EXPENDITURE AMOUNT" in h) and ptd_col is None:
                    ptd_col = idx
                # Use actual Remaining Balance Amount, not Projected Remaining Balance Amount.
                if remaining_col is None and (h == "REMAINING BALANCE AMOUNT" or ("REMAINING BALANCE" in h and "PROJECTED" not in h)):
                    remaining_col = idx

            pivot_found = all(x is not None for x in [row_label_col, month_actual_col, budget_col, ptd_col, remaining_col])
            raw_found = all(x is not None for x in [award_col, task_col, month_actual_col, budget_col, ptd_col, remaining_col])
            found = max(
                sum(x is not None for x in [row_label_col, month_actual_col, budget_col, ptd_col, remaining_col]),
                sum(x is not None for x in [award_col, task_col, month_actual_col, budget_col, ptd_col, remaining_col]),
            )
            if best is None or found > best[0]:
                best = (found, ws.title, header_row, headers[:80])

            if raw_found:
                log(f"[STEP] Using raw/detail sheet '{ws.title}' with header row {header_row}")
                data: Dict[Tuple[str, str], BalanceTaskData] = {}
                max_needed_col = max(award_col, task_col, month_actual_col, budget_col, ptd_col, remaining_col)
                last_award = ""
                last_task = ""

                for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    values = list(values)
                    if len(values) < max_needed_col:
                        values = values + [None] * (max_needed_col - len(values))

                    award = normalize_award(values[award_col - 1])
                    task = normalize_task(values[task_col - 1])
                    if award:
                        last_award = award
                    else:
                        award = last_award
                    if task:
                        last_task = task
                    else:
                        task = last_task

                    if not award or not task:
                        continue
                    if not re.fullmatch(r"[A-Z0-9]{5}", award) or not re.fullmatch(r"\d{1,5}", task):
                        continue

                    key = (award, task)
                    if key not in data:
                        data[key] = BalanceTaskData(award=award, task=task, month_actual=0.0, ptd_actual=0.0, budget=0.0, remaining_balance=0.0)
                    item = data[key]
                    item.month_actual = round(item.month_actual + safe_float(values[month_actual_col - 1]), 2)
                    item.ptd_actual = round(item.ptd_actual + safe_float(values[ptd_col - 1]), 2)
                    item.budget = round(item.budget + safe_float(values[budget_col - 1]), 2)
                    item.remaining_balance = round(item.remaining_balance + safe_float(values[remaining_col - 1]), 2)

                if not data:
                    raise ValueError("Raw/detail source columns were found, but no award/task rows were parsed.")
                log(f"[STEP] Parsed {len(data)} award/task rows from raw/detail balance source workbook")
                return data

            if pivot_found:
                log(f"[STEP] Using pivot sheet '{ws.title}' with header row {header_row}")
                current_award = None
                data: Dict[Tuple[str, str], BalanceTaskData] = {}
                max_needed_col = max(row_label_col, month_actual_col, budget_col, ptd_col, remaining_col)

                for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    values = list(values)
                    if len(values) < max_needed_col:
                        values = values + [None] * (max_needed_col - len(values))
                    label = normalize_text(values[row_label_col - 1])
                    if not label or label.upper() in {"GRAND TOTAL", "ROWS 1 - 90 (ALL ROWS)"}:
                        continue
                    if re.fullmatch(r"[A-Za-z0-9]{5}", label):
                        current_award = normalize_award(label)
                        continue
                    if current_award and re.fullmatch(r"\d{1,5}", label):
                        task = normalize_task(label)
                        data[(current_award, task)] = BalanceTaskData(
                            award=current_award,
                            task=task,
                            month_actual=safe_float(values[month_actual_col - 1]),
                            ptd_actual=safe_float(values[ptd_col - 1]),
                            budget=safe_float(values[budget_col - 1]),
                            remaining_balance=safe_float(values[remaining_col - 1]),
                        )
                if not data:
                    raise ValueError("Pivot source columns were found, but no award/task rows were parsed.")
                log(f"[STEP] Parsed {len(data)} award/task rows from pivot balance source workbook")
                return data

    if best:
        raise KeyError(
            "Balance source columns were not found automatically. "
            f"Best guess: sheet '{best[1]}', header row {best[2]}, found {best[0]} required columns. "
            f"Headers seen: {best[3]}"
        )
    raise KeyError("Balance source columns were not found automatically.")

def parse_award_task_from_name(name: str) -> Tuple[Optional[str], Optional[str]]:
    stem = Path(name).stem.upper()
    matches = list(re.finditer(r"([A-Z0-9]{5})[^A-Z0-9]+(\d{1,5})(?!\d)", stem))
    if matches:
        m = matches[-1]
        return m.group(1), m.group(2)
    tokens = re.findall(r"[A-Z0-9]+", stem)
    for i in range(len(tokens) - 1):
        if re.fullmatch(r"[A-Z0-9]{5}", tokens[i]) and re.fullmatch(r"\d{1,5}", tokens[i + 1]):
            return tokens[i], tokens[i + 1]
    return None, None


def parse_award_sum_from_name(name: str) -> Optional[str]:
    stem = Path(name).stem.upper()
    m = re.search(r"([A-Z0-9]{5})[^A-Z0-9]*SUM\b", stem)
    if m:
        return m.group(1)
    tokens = re.findall(r"[A-Z0-9]+", stem)
    if tokens and tokens[-1] == "SUM":
        for tok in reversed(tokens[:-1]):
            if re.fullmatch(r"[A-Z0-9]{5}", tok):
                return tok
    return None


def is_sum_file(path: Path) -> bool:
    return parse_award_sum_from_name(path.name) is not None


def resolve_shortcut(path: Path) -> Optional[Path]:
    if path.suffix.lower() != ".lnk":
        return None
    if os.name == "nt":
        ps_script = (
            "$s=(New-Object -ComObject WScript.Shell).CreateShortcut('"
            + str(path).replace("'", "''")
            + "'); if ($s.TargetPath) { Write-Output $s.TargetPath }"
        )
        try:
            result = subprocess.run(
                ["powershell", "-NoProfile", "-Command", ps_script],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )
            target = result.stdout.strip()
            return Path(target) if target else None
        except Exception:
            return None
    return None


def discover_files(inputs: Iterable[str], recursive: bool) -> List[WorkItem]:
    items: List[WorkItem] = []

    def add_file(p: Path):
        if p.name.startswith("~$"):
            return
        if p.name.lower() in {"balance_fill_report.xlsx", "fill_report.csv", "commit_fill_report.xlsx", "commit_fill_report.csv"}:
            return
        if p.suffix.lower() in EXCEL_EXTS:
            items.append(WorkItem(input_path=p, workbook_path=p))
        elif p.suffix.lower() == ".lnk":
            target = resolve_shortcut(p)
            if target and target.suffix.lower() in EXCEL_EXTS:
                log(f"[STEP] Resolved shortcut: {p} -> {target}")
                items.append(WorkItem(input_path=p, workbook_path=target))
            else:
                log(f"[STEP] Skipping shortcut with no Excel target: {p}")

    for raw in inputs:
        p = Path(strip_quotes(raw))
        if p.is_dir():
            iterator = p.rglob("*") if recursive else p.glob("*")
            for child in iterator:
                if child.is_file() and child.suffix.lower() in INPUT_EXTS:
                    add_file(child)
        elif p.is_file() and p.suffix.lower() in INPUT_EXTS:
            add_file(p)

    seen = set()
    unique: List[WorkItem] = []
    for item in items:
        key = (str(item.input_path.resolve()), str(item.workbook_path.resolve()))
        if key not in seen:
            seen.add(key)
            unique.append(item)
    return unique


def excel_col(n: int) -> str:
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def excel_com_available() -> bool:
    if os.name != "nt":
        return False
    try:
        import win32com.client  # type: ignore
        return True
    except Exception:
        return False


def get_excel_app():
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except Exception as exc:
        raise RuntimeError("Microsoft Excel + pywin32 are required for validated in-place editing. Run: python -m pip install pywin32") from exc
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    return excel, pythoncom


def normalize_cell_text(value) -> str:
    return normalize_text(value).upper().replace(".", "")


def locate_cells_excel(wb, year: int, month: int) -> Tuple[object, str, str, str, Optional[str], Optional[str]]:
    for ws in wb.Worksheets:
        used = ws.UsedRange
        max_row = min(used.Rows.Count, 80)
        max_col = used.Columns.Count

        current_row = None
        unexpended_row = None
        budget_row = None

        for r in range(1, max_row + 1):
            for c in range(1, min(max_col, 20) + 1):
                text = normalize_cell_text(ws.Cells(r, c).Value)

                if text == "CURRENT MO EXPENSE":
                    current_row = r

                elif text == "UNEXPENDED BALANCE":
                    unexpended_row = r

                elif (
                    text in {"BUDGET", "BUDGET AMOUNT", "TOTAL BUDGET", "AUTHORIZED BUDGET"}
                    or ("BUDGET" in text and "PROJECTED" not in text)
                ):
                    if budget_row is None:
                        budget_row = r

        if current_row is None or unexpended_row is None:
            continue

        header_row = None
        current_col = None

        for r in range(1, min(max_row, 30) + 1):
            for c in range(1, max_col + 1):
                v = ws.Cells(r, c).Value
                matched = False

                if hasattr(v, "year") and hasattr(v, "month"):
                    matched = (v.year == year and v.month == month)
                else:
                    matched = month_matches(v, year, month)

                if matched:
                    header_row = r
                    current_col = c
                    break

            if current_col:
                break

        if current_col is None:
            continue

        style_source = None
        for c in range(current_col - 1, 0, -1):
            v = ws.Cells(header_row, c).Value
            if is_any_month_header(v):
                style_source = f"{excel_col(c)}{current_row}"
                break

        budget_cell = f"{excel_col(current_col)}{budget_row}" if budget_row else None

        return (
            ws,
            f"{excel_col(current_col)}{current_row}",
            f"{excel_col(current_col)}{unexpended_row}",
            f"{excel_col(current_col)}{header_row}",
            style_source,
            budget_cell,
        )

    raise ValueError(
        "Could not find worksheet containing CURRENT MO. EXPENSE, "
        "UNEXPENDED BALANCE, and requested month column."
    )


def extract_award_task_from_workbook_excel(wb) -> Tuple[Optional[str], Optional[str]]:
    for ws in wb.Worksheets:
        used = ws.UsedRange
        max_row = min(used.Rows.Count, 30)
        max_col = min(used.Columns.Count, 12)
        found_award = None
        found_task = None
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                text = normalize_text(ws.Cells(r, c).Value)
                if normalize_cell_text(text) == "AWARD":
                    found_award = normalize_award(ws.Cells(r, c + 1).Value)
                if normalize_cell_text(text) == "TASK":
                    found_task = normalize_task(ws.Cells(r, c + 1).Value)
                if found_award and found_task:
                    return found_award, found_task
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                text = normalize_text(ws.Cells(r, c).Value)
                if not found_award and re.fullmatch(r"[A-Za-z0-9]{5}", text):
                    found_award = normalize_award(text)
                if not found_task and re.fullmatch(r"\d{1,5}", text):
                    found_task = normalize_task(text)
                if found_award and found_task:
                    return found_award, found_task
    return None, None


def identify_award_task(workbook_path: Path, identity_path: Path) -> Tuple[str, str, str]:
    award, task = parse_award_task_from_name(identity_path.name)
    source = []
    if award:
        source.append("award from filename")
    if task:
        source.append("task from filename")
    if not (award and task):
        alt_award, alt_task = parse_award_task_from_name(workbook_path.name)
        award = award or alt_award
        task = task or alt_task
        if alt_award and "award from filename" not in source:
            source.append("award from workbook filename")
        if alt_task and "task from filename" not in source:
            source.append("task from workbook filename")
    if award and task:
        return award, task, ", ".join(source)

    excel, pythoncom = get_excel_app()
    wb = None
    try:
        wb = excel.Workbooks.Open(str(workbook_path), UpdateLinks=False, ReadOnly=True)
        wb_award, wb_task = extract_award_task_from_workbook_excel(wb)
        award = award or wb_award
        task = task or wb_task
        if wb_award:
            source.append("award from workbook")
        if wb_task:
            source.append("task from workbook")
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()
        pythoncom.CoUninitialize()
    if not (award and task):
        raise ValueError(f"Could not determine award/task for {identity_path.name} / {workbook_path.name}.")
    return award, task, ", ".join(source) if source else "unknown"


def process_excel_validated(
    path: Path,
    amount: float,
    expected_balance: float,
    expected_budget: Optional[float],
    year: int,
    month: int,
    recalc: bool,
) -> Dict[str, object]:
    """
    Fill CURRENT MO. EXPENSE first and validate UNEXPENDED BALANCE.

    If the first validation fails, try updating the matching month BUDGET cell.
    - If changing BUDGET makes validation pass, keep CURRENT MO. EXPENSE and BUDGET.
    - If validation still fails, restore the original CURRENT MO. EXPENSE and BUDGET.
    """
    excel, pythoncom = get_excel_app()
    wb = None

    try:
        wb = excel.Workbooks.Open(str(path), UpdateLinks=False, ReadOnly=False)

        (
            ws,
            current_cell,
            unexpended_cell,
            header_cell,
            style_source,
            budget_cell,
        ) = locate_cells_excel(wb, year, month)

        current_rng = ws.Range(current_cell)
        budget_rng = ws.Range(budget_cell) if budget_cell else None

        # Store formulas, not just values, so a failed validation can restore cells exactly.
        old_current_value = current_rng.Value
        old_current_formula = current_rng.Formula
        old_budget_value = budget_rng.Value if budget_rng is not None else None
        old_budget_formula = budget_rng.Formula if budget_rng is not None else None

        def recalculate() -> None:
            if recalc:
                excel.CalculateFullRebuild()
            else:
                excel.Calculate()

        def read_unexpended_balance() -> Optional[float]:
            raw = ws.Range(unexpended_cell).Value
            try:
                return float(raw)
            except Exception:
                return None

        def restore_original_values() -> None:
            current_rng.Formula = old_current_formula
            if budget_rng is not None:
                budget_rng.Formula = old_budget_formula
            recalculate()

        def apply_current_cell_format() -> None:
            # Apply formatting only when a change is actually kept.
            if style_source:
                ws.Range(style_source).Copy()
                current_rng.PasteSpecial(Paste=-4122)  # xlPasteFormats
                excel.CutCopyMode = False

        # First attempt: fill only CURRENT MO. EXPENSE.
        current_rng.Value = float(amount)
        recalculate()

        actual_balance = read_unexpended_balance()

        if actual_balance is None:
            restore_original_values()
            wb.Save()
            return {
                "status": "FAIL",
                "filled": False,
                "budget_updated": False,
                "sheet": ws.Name,
                "month_cell": current_cell,
                "budget_cell": budget_cell or "",
                "month_header_cell": header_cell,
                "style_source_cell": style_source or "",
                "workbook_unexpended_balance": "",
                "budget_trial_unexpended_balance": "",
                "old_workbook_budget": old_budget_value if budget_rng is not None else "",
                "new_workbook_budget": "",
                "notes": (
                    f"Could not read numeric UNEXPENDED BALANCE at {unexpended_cell}; "
                    "CURRENT MO. EXPENSE and BUDGET were restored to their original values."
                ),
            }

        diff = round(actual_balance - expected_balance, 2)

        # Normal validation passed. No budget change needed.
        if abs(diff) <= TOLERANCE:
            apply_current_cell_format()
            wb.Save()
            return {
                "status": "OK",
                "filled": True,
                "budget_updated": False,
                "sheet": ws.Name,
                "month_cell": current_cell,
                "budget_cell": budget_cell or "",
                "month_header_cell": header_cell,
                "style_source_cell": style_source or "",
                "workbook_unexpended_balance": actual_balance,
                "budget_trial_unexpended_balance": "",
                "old_workbook_budget": old_budget_value if budget_rng is not None else "",
                "new_workbook_budget": "",
                "notes": "Validated successfully; CURRENT MO. EXPENSE filled. Budget was not changed.",
            }

        # If normal validation failed, try updating BUDGET.
        if expected_budget is not None and budget_rng is not None:
            budget_rng.Value = float(expected_budget)
            recalculate()

            budget_trial_balance = read_unexpended_balance()

            if budget_trial_balance is not None:
                budget_trial_diff = round(budget_trial_balance - expected_balance, 2)

                # Budget update fixed the validation.
                if abs(budget_trial_diff) <= TOLERANCE:
                    apply_current_cell_format()
                    wb.Save()
                    return {
                        "status": "OK",
                        "filled": True,
                        "budget_updated": True,
                        "sheet": ws.Name,
                        "month_cell": current_cell,
                        "budget_cell": budget_cell or "",
                        "month_header_cell": header_cell,
                        "style_source_cell": style_source or "",
                        "workbook_unexpended_balance": budget_trial_balance,
                        "budget_trial_unexpended_balance": budget_trial_balance,
                        "old_workbook_budget": old_budget_value if budget_rng is not None else "",
                        "new_workbook_budget": expected_budget,
                        "notes": (
                            f"Initial validation failed: workbook UNEXPENDED BALANCE "
                            f"{actual_balance:,.2f} did not match source remaining balance "
                            f"{expected_balance:,.2f}. Budget was updated from "
                            f"{safe_float(old_budget_value):,.2f} to {expected_budget:,.2f}; "
                            "second validation passed. CURRENT MO. EXPENSE and BUDGET were both kept."
                        ),
                    }

            # Budget update did not fix validation, or trial balance was unreadable.
            restore_original_values()
            wb.Save()

            return {
                "status": "CHECK",
                "filled": False,
                "budget_updated": False,
                "sheet": ws.Name,
                "month_cell": current_cell,
                "budget_cell": budget_cell or "",
                "month_header_cell": header_cell,
                "style_source_cell": style_source or "",
                "workbook_unexpended_balance": actual_balance,
                "budget_trial_unexpended_balance": budget_trial_balance if budget_trial_balance is not None else "",
                "old_workbook_budget": old_budget_value if budget_rng is not None else "",
                "new_workbook_budget": "",
                "notes": (
                    f"Initial validation failed: workbook UNEXPENDED BALANCE "
                    f"{actual_balance:,.2f} did not match source remaining balance "
                    f"{expected_balance:,.2f}. Tried updating BUDGET to "
                    f"{expected_budget:,.2f}, but validation still failed. "
                    "CURRENT MO. EXPENSE and BUDGET were restored to their original values."
                ),
            }

        # No budget cell found, or no expected budget available.
        restore_original_values()
        wb.Save()

        reason = "No matching BUDGET cell was found" if budget_rng is None else "No expected budget was available from source data"

        return {
            "status": "CHECK",
            "filled": False,
            "budget_updated": False,
            "sheet": ws.Name,
            "month_cell": current_cell,
            "budget_cell": budget_cell or "",
            "month_header_cell": header_cell,
            "style_source_cell": style_source or "",
            "workbook_unexpended_balance": actual_balance,
            "budget_trial_unexpended_balance": "",
            "old_workbook_budget": old_budget_value if budget_rng is not None else "",
            "new_workbook_budget": "",
            "notes": (
                f"Validation failed: workbook UNEXPENDED BALANCE {actual_balance:,.2f} "
                f"does not match source remaining balance {expected_balance:,.2f}. "
                f"{reason}, so BUDGET was not changed. CURRENT MO. EXPENSE was restored."
            ),
        }

    finally:
        if wb is not None:
            # Any successful/intentional change is saved explicitly with wb.Save().
            # This close call prevents accidental partial edits from being saved after errors.
            wb.Close(SaveChanges=False)

        excel.Quit()
        pythoncom.CoUninitialize()


def build_folder_task_map(files: List[WorkItem]) -> Dict[str, set]:
    tasks_by_award: Dict[str, set] = {}
    for item in files:
        # SUM files do not count as task files.
        names = [item.input_path.name, item.workbook_path.name]
        if any(parse_award_sum_from_name(n) for n in names):
            continue
        award, task = parse_award_task_from_name(item.input_path.name)
        if not (award and task):
            award, task = parse_award_task_from_name(item.workbook_path.name)
        if award and task:
            tasks_by_award.setdefault(award, set()).add(task)
    return tasks_by_award


def detect_sum_awards(files: List[WorkItem]) -> set:
    awards = set()
    for item in files:
        for name in [item.input_path.name, item.workbook_path.name]:
            award = parse_award_sum_from_name(name)
            if award:
                awards.add(award)
                break
    return awards


def write_processing_report(output_dir: str, rows: List[Dict[str, object]]) -> str:
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, "balance_fill_report.xlsx")
    headers = [
        "status", "filled", "file_type", "path", "award", "task", "tasks_found",
        "month_actual", "expected_budget", "expected_remaining_balance",
        "workbook_unexpended_balance", "budget_trial_unexpended_balance",
        "difference", "budget_updated", "old_workbook_budget", "new_workbook_budget",
        "target_sheet", "month_cell", "budget_cell", "month_header_cell", "style_source_cell",
        "task_totals", "notes"
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Processing Report"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    status_col = headers.index("status") + 1
    for r in range(2, ws.max_row + 1):
        status = str(ws.cell(r, status_col).value or "").upper()
        fill = None
        if status == "OK":
            fill = PatternFill("solid", fgColor="E2F0D9")
        elif status == "CHECK":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif status == "FAIL":
            fill = PatternFill("solid", fgColor="FCE4D6")
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill

    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            max_len = max(max_len, len(str(ws.cell(row_idx, col_idx).value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 70)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    log(f"Excel processing report written to: {path}")
    return path


def choose_default_output_dir(inputs: List[str]) -> str:
    for item in inputs:
        p = Path(strip_quotes(item))
        if p.is_dir():
            return str(p)
        if p.is_file():
            return str(p.parent)
    return os.getcwd()


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill balance graph workbooks from Expenditure Summary raw/pivot source.")
    parser.add_argument("--raw", "--pivot", dest="raw", help="Path to Expenditure Summary workbook.")
    parser.add_argument("--inputs", nargs="+", help="One or more target files/folders/shortcuts.")
    parser.add_argument("--month", help="Month label like mar-2026.")
    parser.add_argument("--output-dir", default=None, help="Where to write balance_fill_report.xlsx.")
    parser.add_argument("--include-subfolders", action="store_true", help="Scan folders recursively.")
    parser.add_argument("--recalc", action="store_true", help="Use full Excel recalculation before validation.")
    parser.add_argument("--no-report", action="store_true", help="Do not create processing report.")
    args = parser.parse_args()

    if not args.raw:
        args.raw = strip_quotes(input("Enter FULL path to raw Expenditure Summary Excel: "))
    else:
        args.raw = strip_quotes(args.raw)
    if not args.inputs:
        args.inputs = [strip_quotes(input("Enter FULL path to Input folder/files: "))]
    else:
        args.inputs = [strip_quotes(x) for x in args.inputs]
    if not args.month:
        args.month = strip_quotes(input("Enter month (e.g. mar-2026): "))
    if not args.output_dir:
        args.output_dir = choose_default_output_dir(args.inputs)
    else:
        args.output_dir = strip_quotes(args.output_dir)

    print("\n=== RUNNING WITH ===")
    print(f"Raw source: {args.raw}")
    print(f"Inputs: {args.inputs}")
    print(f"Month:  {args.month}")
    print(f"Output report folder: {args.output_dir}")
    print("=" * 30)

    try:
        year, month, month_header_key, month_display = parse_month_input(args.month)
    except Exception as exc:
        print(f"[FAIL] Month could not be parsed: {exc}")
        return 1

    files = discover_files(args.inputs, recursive=args.include_subfolders)
    log(f"[STEP] Found {len(files)} workbook/shortcut file(s) to process")
    if not files:
        print("No Excel files were found in the provided inputs.")
        return 1

    report_rows: List[Dict[str, object]] = []
    try:
        source_index = read_balance_report(args.raw, month_header_key)
    except Exception as exc:
        reason = f"Balance source workbook could not be read, so no workbooks were updated. Reason: {exc}"
        print(f"[FAIL] {reason}")
        for item in files:
            award, task = parse_award_task_from_name(item.input_path.name)
            file_type = "TASK"
            if is_sum_file(item.input_path) or is_sum_file(item.workbook_path):
                award = parse_award_sum_from_name(item.input_path.name) or parse_award_sum_from_name(item.workbook_path.name)
                task = "SUM"
                file_type = "SUM"
            report_rows.append({"status": "FAIL", "filled": "NO", "file_type": file_type, "path": str(item.workbook_path), "award": award or "", "task": task or "", "notes": reason})
        if not args.no_report:
            write_processing_report(args.output_dir, report_rows)
        return 1

    sum_awards = detect_sum_awards(files)
    tasks_by_award = build_folder_task_map(files)
    if sum_awards:
        log(f"[STEP] SUM workbook mode detected for award(s): {', '.join(sorted(sum_awards))}")
    else:
        log("[STEP] No SUM workbooks detected.")

    needs_review = []
    failures = []

    for item in files:
        path = item.workbook_path
        input_path = item.input_path
        if Path(args.raw).resolve() == path.resolve():
            continue
        try:
            log(f"\n[FILE] {path}")
            if is_sum_file(input_path) or is_sum_file(path):
                award = parse_award_sum_from_name(input_path.name) or parse_award_sum_from_name(path.name)
                if not award:
                    raise ValueError("Could not identify SUM award from filename.")
                folder_tasks = sorted(tasks_by_award.get(award, set()), key=lambda x: int(x) if x.isdigit() else x)
                if not folder_tasks:
                    raise ValueError(f"SUM workbook found for {award}, but no task files for this award were found in the selected folder.")
                missing_from_source = [t for t in folder_tasks if (award, t) not in source_index]
                if missing_from_source:
                    raise KeyError(f"SUM {award}: task file(s) present but missing from source data: {', '.join(missing_from_source)}")

                month_actual = round(sum(source_index[(award, t)].month_actual for t in folder_tasks), 2)
                expected_budget = round(sum(source_index[(award, t)].budget for t in folder_tasks), 2)
                expected_remaining = round(sum(source_index[(award, t)].remaining_balance for t in folder_tasks), 2)
                raw_tasks_for_award = sorted({t for (a, t) in source_index if a == award}, key=lambda x: int(x) if x.isdigit() else x)
                skipped = [t for t in raw_tasks_for_award if t not in folder_tasks]
                result = process_excel_validated(path, month_actual, expected_remaining, expected_budget, year, month, args.recalc)
                diff = ""
                if isinstance(result.get("workbook_unexpended_balance"), (int, float)):
                    diff = f"{round(float(result['workbook_unexpended_balance']) - expected_remaining, 2):.2f}"
                notes = str(result.get("notes", ""))
                if skipped:
                    notes += f" Skipped source task(s) not present in selected folder: {', '.join(skipped)}."
                task_totals = "; ".join(f"{t}:{source_index[(award,t)].month_actual:.2f}" for t in folder_tasks)
                report_rows.append({
                    "status": result["status"], 
                    "filled": "YES" if result["filled"] else "NO", 
                    "file_type": "SUM",
                    "path": str(path), 
                    "award": award, 
                    "task": "SUM", 
                    "tasks_found": len(folder_tasks),
                    "month_actual": f"{month_actual:.2f}",
                    "expected_budget": f"{expected_budget:.2f}",
                    "expected_remaining_balance": f"{expected_remaining:.2f}",
                    "workbook_unexpended_balance": result.get("workbook_unexpended_balance", ""),
                    "budget_trial_unexpended_balance": result.get("budget_trial_unexpended_balance", ""),
                    "difference": diff,
                    "budget_updated": "YES" if result.get("budget_updated") else "NO",
                    "old_workbook_budget": result.get("old_workbook_budget", ""),
                    "new_workbook_budget": result.get("new_workbook_budget", ""),
                    "target_sheet": result.get("sheet", ""),
                    "month_cell": result.get("month_cell", ""),
                    "budget_cell": result.get("budget_cell", ""),
                    "month_header_cell": result.get("month_header_cell", ""),
                    "style_source_cell": result.get("style_source_cell", ""),
                    "task_totals": task_totals, 
                    "notes": notes,
                })
                print(f"[{result['status']}] {path.name} | {award}/SUM | amount {month_actual:,.2f} | filled={result['filled']}")
                if result["status"] != "OK":
                    needs_review.append(f"{award}.SUM")
                continue

            award, task, source_used = identify_award_task(path, input_path)
            if (award, task) not in source_index:
                raise KeyError(f"No source data found for award/task {award}/{task}.")
            pdata = source_index[(award, task)]

            result = process_excel_validated(
                path,
                pdata.month_actual,
                pdata.remaining_balance,
                pdata.budget,
                year,
                month,
                args.recalc,
            )
            diff = ""
            if isinstance(result.get("workbook_unexpended_balance"), (int, float)):
                diff = f"{round(float(result['workbook_unexpended_balance']) - pdata.remaining_balance, 2):.2f}"
            notes = str(result.get("notes", ""))
            if award in sum_awards:
                notes += f" SUM workbook exists for award {award}; this task workbook is still validated and filled independently."
            report_rows.append({
                "status": result["status"], 
                "filled": "YES" if result["filled"] else "NO", 
                "file_type": "TASK",
                "path": str(path), 
                "award": award, 
                "task": task, 
                "tasks_found": "",
                "month_actual": f"{pdata.month_actual:.2f}",
                "expected_budget": f"{pdata.budget:.2f}",
                "expected_remaining_balance": f"{pdata.remaining_balance:.2f}",
                "workbook_unexpended_balance": result.get("workbook_unexpended_balance", ""),
                "budget_trial_unexpended_balance": result.get("budget_trial_unexpended_balance", ""),
                "difference": diff,
                "budget_updated": "YES" if result.get("budget_updated") else "NO",
                "old_workbook_budget": result.get("old_workbook_budget", ""),
                "new_workbook_budget": result.get("new_workbook_budget", ""),
                "target_sheet": result.get("sheet", ""),
                "month_cell": result.get("month_cell", ""),
                "budget_cell": result.get("budget_cell", ""),
                "month_header_cell": result.get("month_header_cell", ""),
                "style_source_cell": result.get("style_source_cell", ""),
                "task_totals": "", 
                "notes": notes,
            })
            print(f"[{result['status']}] {path.name} | {award}/{task} | amount {pdata.month_actual:,.2f} | filled={result['filled']}")
            if result["status"] != "OK":
                needs_review.append(f"{award}.{task}")
        except Exception as exc:
            award, task = parse_award_task_from_name(input_path.name)
            file_type = "TASK"
            if is_sum_file(input_path) or is_sum_file(path):
                award = parse_award_sum_from_name(input_path.name) or parse_award_sum_from_name(path.name)
                task = "SUM"
                file_type = "SUM"
            failures.append(f"{path}: {exc}")
            if award and task:
                needs_review.append(f"{award}.{task}")
            report_rows.append({"status": "FAIL", "filled": "NO", "file_type": file_type, "path": str(path), "award": award or "", "task": task or "", "notes": str(exc)})
            print(f"[FAIL] {path}: {exc}")

    if not args.no_report:
        write_processing_report(args.output_dir, report_rows)

    print("\nAward.task items needing review:")
    if needs_review:
        for item in sorted(set(needs_review)):
            print(item)
    else:
        print("None")

    if failures:
        print("\nFailures:")
        for item in failures:
            print(f"  - {item}")
        return 2
    if any(str(r.get("status", "")).upper() != "OK" for r in report_rows):
        return 3
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
