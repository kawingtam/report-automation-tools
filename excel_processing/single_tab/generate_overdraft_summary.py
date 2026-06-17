"""
PTA Summary OD Report Generator

This version reads directly from the SOURCE/RAW table instead of relying on the pivot table.

It automatically searches worksheets for columns such as:
- Project Owner Full Name
- Award / Award Number
- Task / Task Number
- Remaining Balance Amount

Then it creates:
PTA SUMMARY_YYYY-MM-DD_HHMMSS.xlsx

The report lists award-task rows where Remaining Balance Amount is negative.
"""

from __future__ import annotations

import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from typing import Any, Optional

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: The Python package 'openpyxl' is required.")
    print("Install it with: py -m pip install openpyxl")
    input("Press Enter to exit...")
    sys.exit(1)


TASK_RE = re.compile(r"^\d+(?:\.0)?$")


def clean_text(value: Any) -> str:
    """Convert Excel cell value to a clean string."""
    if value is None:
        return ""

    text = str(value).strip()

    if text.endswith(".0") and TASK_RE.match(text):
        text = text[:-2]

    return text


def normalize_header(value: Any) -> str:
    """Normalize headers for flexible matching."""
    text = clean_text(value).lower()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def to_number(value: Any) -> Optional[float]:
    """Convert Excel values like -1904.53 or ($1,904.53) to float."""
    if value is None or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    if not text:
        return None

    negative = False

    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    text = text.replace("$", "").replace(",", "").replace(" ", "")

    try:
        num = float(text)
        return -num if negative else num
    except ValueError:
        return None


def format_currency(amount: float) -> str:
    """Format OD amount as ($1,904.53)."""
    amount = abs(float(amount))
    return f"(${amount:,.2f})"


def find_column(headers: list[str], exact_names: list[str], contains_names: list[str]) -> Optional[int]:
    """
    Find a column by exact header first, then by contains match.
    Returns 1-based Excel column number.
    """
    normalized = [normalize_header(h) for h in headers]

    for idx, header in enumerate(normalized, start=1):
        if header in exact_names:
            return idx

    for idx, header in enumerate(normalized, start=1):
        for keyword in contains_names:
            if keyword in header:
                return idx

    return None


def find_source_table(wb):
    """
    Search all worksheets for a source/raw table.

    Required:
    - Project Owner Full Name
    - Award
    - Task
    - Remaining Balance Amount

    Returns:
        ws, header_row, owner_col, award_col, task_col, remaining_col
    """
    candidates = []

    for ws in wb.worksheets:
        for row in range(1, min(ws.max_row, 80) + 1):
            headers = [
                clean_text(ws.cell(row=row, column=col).value)
                for col in range(1, ws.max_column + 1)
            ]

            owner_col = find_column(
                headers,
                exact_names=[
                    "project owner full name",
                ],
                contains_names=[
                    "project owner full name",
                    "project owner",
                    "owner full name",
                ],
            )

            award_col = find_column(
                headers,
                exact_names=[
                    "award",
                    "award number",
                    "award code",
                    "award id",
                ],
                contains_names=[
                    "award number",
                    "award code",
                    "award id",
                    "sponsored award",
                    "award",
                ],
            )

            task_col = find_column(
                headers,
                exact_names=[
                    "task",
                    "task number",
                    "task code",
                    "task id",
                ],
                contains_names=[
                    "task number",
                    "task code",
                    "task id",
                    "task",
                ],
            )

            remaining_col = find_column(
                headers,
                exact_names=[
                    "remaining balance amount",
                    "remaining balance",
                    "sum of remaining balance amount",
                ],
                contains_names=[
                    "remaining balance amount",
                    "remaining balance",
                ],
            )

            # We only accept a real source table if all required columns are found.
            if owner_col and award_col and task_col and remaining_col:
                non_blank_rows = count_data_rows(
                    ws,
                    header_row=row,
                    required_cols=[owner_col, award_col, task_col, remaining_col],
                )

                candidates.append(
                    {
                        "ws": ws,
                        "header_row": row,
                        "owner_col": owner_col,
                        "award_col": award_col,
                        "task_col": task_col,
                        "remaining_col": remaining_col,
                        "non_blank_rows": non_blank_rows,
                    }
                )

    if not candidates:
        raise ValueError(
            "Could not find a source/raw table with the required columns:\n"
            "- Project Owner Full Name\n"
            "- Award\n"
            "- Task\n"
            "- Remaining Balance Amount\n\n"
            "Make sure the workbook has the raw/source data as a normal worksheet, "
            "not only a pivot table."
        )

    # Pick the source table with the most usable data rows.
    candidates.sort(key=lambda x: x["non_blank_rows"], reverse=True)
    best = candidates[0]

    return (
        best["ws"],
        best["header_row"],
        best["owner_col"],
        best["award_col"],
        best["task_col"],
        best["remaining_col"],
    )


def count_data_rows(ws, header_row: int, required_cols: list[int]) -> int:
    """Count rows that appear to have source data."""
    count = 0

    for row in range(header_row + 1, ws.max_row + 1):
        values = [
            clean_text(ws.cell(row=row, column=col).value)
            for col in required_cols
        ]

        if any(values):
            count += 1

    return count


def build_summary_rows_from_source(
    ws,
    header_row: int,
    owner_col: int,
    award_col: int,
    task_col: int,
    remaining_col: int,
) -> list[list[Any]]:
    """
    Build OD summary rows directly from the source table.

    Groups by:
    - Project Owner Full Name
    - Award
    - Task

    Then sums Remaining Balance Amount and reports groups where the total is negative.
    """
    totals = defaultdict(float)

    for row in range(header_row + 1, ws.max_row + 1):
        owner = clean_text(ws.cell(row=row, column=owner_col).value)
        award = clean_text(ws.cell(row=row, column=award_col).value).upper()
        task = clean_text(ws.cell(row=row, column=task_col).value)
        remaining_balance = to_number(ws.cell(row=row, column=remaining_col).value)

        if not award or not task:
            continue

        if remaining_balance is None:
            continue

        # Clean task number like 601.0 -> 601.
        if task.endswith(".0") and TASK_RE.match(task):
            task = task[:-2]

        totals[(owner, award, task)] += remaining_balance

    results = []

    for (owner, award, task), remaining_balance in totals.items():
        if remaining_balance < 0:
            award_task = f"{award}-{task}"

            if owner:
                sentence = (
                    f"{award_task} There is an OD "
                    f"{format_currency(remaining_balance)}."
                )
            else:
                sentence = f"{award_task} There is an OD {format_currency(remaining_balance)}."

            results.append([
                owner,
                award,
                task,
                remaining_balance,
                sentence,
            ])

    def sort_key(row_data):
        owner, award, task, *_ = row_data

        try:
            task_sort = int(task)
        except ValueError:
            task_sort = task

        return str(owner), str(award), task_sort

    results.sort(key=sort_key)

    return results


def autosize_columns(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0

        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 75)


def create_output_workbook(
    source_path: str,
    source_sheet: str,
    owner_header: str,
    award_header: str,
    task_header: str,
    remaining_header: str,
    rows: list[list[Any]],
) -> str:
    source_dir = os.path.dirname(os.path.abspath(source_path))
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_path = os.path.join(source_dir, f"PTA SUMMARY_{timestamp}.xlsx")

    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "PTA Summary"

    ws["A1"] = "PTA Overdraft Summary"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:E1")

    ws["A2"] = "Source file"
    ws["B2"] = os.path.basename(source_path)

    ws["A3"] = "Source sheet"
    ws["B3"] = source_sheet

    ws["A4"] = "Owner column used"
    ws["B4"] = owner_header

    ws["A5"] = "Award column used"
    ws["B5"] = award_header

    ws["A6"] = "Task column used"
    ws["B6"] = task_header

    ws["A7"] = "OD column used"
    ws["B7"] = remaining_header

    ws["A8"] = "Generated"
    ws["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    headers = [
        "Project Owner Full Name",
        "Award",
        "Task",
        "Remaining Balance",
        "Summary",
    ]

    ws.append([])
    ws.append(headers)
    header_row = 10

    for row_data in rows:
        ws.append(row_data)

    if not rows:
        ws.append(["", "", "", "", "No overdraft rows found."])

    # Style metadata.
    meta_bold = Font(bold=True)

    for r in range(2, 9):
        ws.cell(r, 1).font = meta_bold

    # Style table.
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=5,
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        row[3].number_format = '$#,##0.00;($#,##0.00);$0.00'

    ws.freeze_panes = "A11"

    autosize_columns(ws)

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 28)
    ws.column_dimensions["E"].width = 75

    out_wb.save(output_path)

    return output_path


def main() -> None:
    print("=== PTA SUMMARY OD REPORT TOOL ===")
    print("Reading from source/raw table, not pivot table.")

    if len(sys.argv) >= 2:
        input_path = sys.argv[1].strip().strip('"')
    else:
        input_path = input("Enter FULL path to Excel workbook: ").strip().strip('"')

    if not input_path:
        print("No file path entered.")
        input("Press Enter to exit...")
        return

    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}")
        input("Press Enter to exit...")
        return

    try:
        wb = load_workbook(input_path, data_only=True)

        (
            ws,
            header_row,
            owner_col,
            award_col,
            task_col,
            remaining_col,
        ) = find_source_table(wb)

        owner_header = clean_text(ws.cell(row=header_row, column=owner_col).value)
        award_header = clean_text(ws.cell(row=header_row, column=award_col).value)
        task_header = clean_text(ws.cell(row=header_row, column=task_col).value)
        remaining_header = clean_text(ws.cell(row=header_row, column=remaining_col).value)

        rows = build_summary_rows_from_source(
            ws,
            header_row,
            owner_col,
            award_col,
            task_col,
            remaining_col,
        )

        output_path = create_output_workbook(
            input_path,
            ws.title,
            owner_header,
            award_header,
            task_header,
            remaining_header,
            rows,
        )

        print("\nSource table found:")
        print(f"Sheet: {ws.title}")
        print(f"Header row: {header_row}")
        print(f"Project Owner column: {owner_header}")
        print(f"Award column: {award_header}")
        print(f"Task column: {task_header}")
        print(f"Remaining Balance column: {remaining_header}")

        print(f"\nDone. Found {len(rows)} overdraft award-task row(s).")
        print(f"Output saved here:\n{output_path}")

    except Exception as exc:
        print("\nERROR:")
        print(exc)

    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()