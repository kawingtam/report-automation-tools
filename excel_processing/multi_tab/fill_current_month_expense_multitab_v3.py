#!/usr/bin/env python3
"""
Fill CURRENT MONTH EXPENSE for multi-PTA workbook(s) from an Expenditure Summary raw export.

Folder + shortcut-enabled version:
- Target input can be:
    1. one target workbook,
    2. one Windows .lnk shortcut to a workbook,
    3. one folder containing many target workbooks,
    4. one folder containing many .lnk shortcuts to target workbooks.
- For each target workbook, the script goes through every worksheet/tab.
- For each worksheet/tab, reads:
    B3 = Award
    B5 = Task
- Skips tabs that do not have a usable Award/Task or do not contain the target row/month column.
- Finds the row in column A labeled CURRENT MONTH EXPENSE / CURRENT MO. EXPENSE.
- Finds the target workbook month column matching the month entered by the user, such as may-2026.
- Fills the intersection cell with the aggregated raw-source column:
    <MMM-YYYY> Actual Expenditure Amount
- Target month and source month can be different by using --source-month.
- Aggregates source raw rows by Award Number + Task Number.
- If no matching source rows are found for a worksheet's Award/Task, writes 0.
- Validates/report-checks the workbook EXPENSE row for the same month against:
    FTD or PTD Actual Expenditure Amount
- Writes one Excel processing report for all processed workbooks.

Notes:
- The script edits target workbook(s) in place. No backup files are created by this program.
- .xlsx and .xlsm target/source files use openpyxl.
- .xls files require Microsoft Excel + pywin32 on Windows.
- .lnk shortcuts are resolved to their real Excel workbook targets before processing.
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import subprocess
import sys
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TOLERANCE = 0.01
EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}
OPENPYXL_EXTS = {".xlsx", ".xlsm"}
SHORTCUT_EXT = ".lnk"


@dataclass
class SourcePTAData:
    award: str
    task: str
    month_actual: float = 0.0
    ftd_ptd_actual: float = 0.0
    row_count: int = 0
    project_number: str = ""
    project_owner_full_name: str = ""
    task_name: str = ""
    award_name: str = ""


@dataclass
class SheetResult:
    status: str
    target_workbook: str
    sheet: str
    shortcut_path: str = ""
    award: str = ""
    task: str = ""
    source_rows: int = 0
    source_month_actual: float = 0.0
    source_ftd_ptd_actual: float = 0.0
    written_value: Optional[float] = None
    previous_value: object = ""
    workbook_expense_value: Optional[float] = None
    difference_current_month_vs_source_month: Optional[float] = None
    difference_expense_vs_ftd_ptd: Optional[float] = None
    month_cell: str = ""
    expense_cell: str = ""
    month_header_cell: str = ""
    project_number: str = ""
    project_owner_full_name: str = ""
    task_name: str = ""
    award_name: str = ""
    notes: str = ""


@dataclass
class TargetWorkbook:
    actual_path: Path
    shortcut_path: Optional[Path] = None


def log(msg: str) -> None:
    print(msg)


def strip_quotes(value) -> str:
    return str(value or "").strip().strip('"').strip("'")


def clean_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_header(value) -> str:
    return clean_header(value).upper()


def normalize_label(value) -> str:
    text = str(value or "").strip().upper()
    text = text.replace(".", " ")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_award(value) -> str:
    """
    Keep the full award token. Do not assume award length is exactly 5.
    """
    text = str(value or "").strip().upper()

    if not text:
        return ""

    if re.fullmatch(r"\d+\.0", text):
        text = text.replace(".0", "")

    tokens = re.findall(r"[A-Z0-9]+", text)

    if not tokens:
        return ""

    skip = {"AWARD", "NUMBER", "NO", "TASK", "PROJECT", "LEGACY", "NA", "NONE"}
    usable = [tok for tok in tokens if tok not in skip]

    if not usable:
        return ""

    # If there are multiple tokens, the actual award is usually the last useful token.
    return usable[-1]


def normalize_task(value) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    if isinstance(value, int):
        return str(value)

    text = str(value).strip()

    if not text:
        return ""

    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]

    m = re.search(r"\d+", text)

    return m.group(0) if m else ""


def safe_float(value) -> float:
    if value is None or value == "":
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    if not text:
        return 0.0

    # Handle accounting-style negatives: ($1,234.56)
    neg = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "")

    try:
        number = float(text)
        return -number if neg else number
    except Exception:
        return 0.0


def parse_month_input(month_text: str) -> Tuple[int, int, str, str]:
    """
    Return:
        year,
        month,
        source_month_key, example MAY-2026,
        display_month, example may-2026
    """
    cleaned = strip_quotes(month_text).replace("_", "-").replace("/", "-")
    cleaned = re.sub(r"\s+", " ", cleaned)

    formats = [
        "%b-%Y", "%b-%y", "%B-%Y", "%B-%y",
        "%b %Y", "%b %y", "%B %Y", "%B %y",
        "%Y-%m", "%Y %m", "%m-%Y", "%m-%y",
    ]

    for fmt in formats:
        try:
            parsed = dt.datetime.strptime(cleaned, fmt)
            return (
                parsed.year,
                parsed.month,
                parsed.strftime("%b-%Y").upper(),
                parsed.strftime("%b-%Y").lower(),
            )
        except ValueError:
            continue

    raise ValueError(f"Could not parse month '{month_text}'. Try values like may-2026 or 2026-05.")


def month_matches(value, year: int, month: int) -> bool:
    if isinstance(value, (dt.datetime, dt.date)):
        return value.year == year and value.month == month

    text = str(value or "").strip()

    if not text:
        return False

    text = text.replace("/", "-")
    text = re.sub(r"\s+", " ", text)

    formats = ["%b-%Y", "%b-%y", "%B-%Y", "%B-%y", "%Y-%m", "%m-%Y", "%m-%y"]

    for fmt in formats:
        try:
            parsed = dt.datetime.strptime(text, fmt)
            return parsed.year == year and parsed.month == month
        except ValueError:
            continue

    return False


def is_any_month_header(value) -> bool:
    if isinstance(value, (dt.datetime, dt.date)):
        return True

    text = str(value or "").strip().replace("/", "-")

    if not text:
        return False

    formats = ["%b-%Y", "%b-%y", "%B-%Y", "%B-%y", "%Y-%m", "%m-%Y", "%m-%y"]

    for fmt in formats:
        try:
            dt.datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue

    return False


def current_month_expense_label(value) -> bool:
    text = normalize_label(value)

    return text in {
        "CURRENT MONTH EXPENSE",
        "CURRENT MONTH EXPENSES",
        "CURRENT MO EXPENSE",
        "CURRENT MO EXPENSES",
        "CURRENT MON EXPENSE",
        "CURRENT MON EXPENSES",
    }


def expense_label(value) -> bool:
    """
    Return True for the cumulative EXPENSE row in the target workbook.

    This intentionally does NOT match CURRENT MONTH EXPENSE.
    Validation compares this EXPENSE row to the source FTD/PTD amount.
    """
    text = normalize_label(value)

    return text in {"EXPENSE", "EXPENSES"}


def format_money(value: Optional[float]) -> str:
    if value is None:
        return "not found"

    return f"{value:,.2f}"




def same_file(path1: Path, path2: Path) -> bool:
    try:
        return path1.resolve().samefile(path2.resolve())
    except Exception:
        return os.path.normcase(os.path.abspath(str(path1))) == os.path.normcase(os.path.abspath(str(path2)))


def resolve_windows_shortcut(path: Path) -> Path:
    """
    Resolve a Windows .lnk shortcut to its real target file.

    If path is not a .lnk shortcut, returns path unchanged.
    """
    if path.suffix.lower() != SHORTCUT_EXT:
        return path

    if os.name != "nt":
        raise RuntimeError(f"Cannot resolve Windows shortcut on this system: {path}")

    # First try pywin32.
    try:
        import win32com.client  # type: ignore

        shell = win32com.client.Dispatch("WScript.Shell")
        shortcut = shell.CreateShortcut(str(path))
        target = strip_quotes(shortcut.TargetPath)

        if target:
            return Path(target).expanduser()

    except Exception:
        pass

    # Fallback to PowerShell.
    escaped = str(path).replace("'", "''")

    cmd = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-Command",
        (
            f"$s=(New-Object -COM WScript.Shell).CreateShortcut('{escaped}'); "
            f"[Console]::WriteLine($s.TargetPath)"
        ),
    ]

    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=20,
    )

    target = strip_quotes(result.stdout)

    if result.returncode == 0 and target:
        return Path(target).expanduser()

    raise RuntimeError(f"Could not resolve shortcut: {path}")


def should_skip_target_candidate(path: Path) -> bool:
    """
    Skip temporary files, reports, and backups before resolving/processing.

    Allows:
    - real Excel files
    - .lnk shortcut files
    """
    name = path.name.lower()
    stem = path.stem.lower()
    suffix = path.suffix.lower()

    if suffix not in EXCEL_EXTS and suffix != SHORTCUT_EXT:
        return True

    # Skip temporary Excel lock files.
    if name.startswith("~$"):
        return True

    # Skip reports created by this script.
    if stem.startswith("current_month_expense_fill_report"):
        return True

    # Skip backups created by this script.
    if re.search(r"_backup_\d{8}_\d{6}$", stem):
        return True

    return False


def should_skip_resolved_workbook(path: Path, source_path: Path) -> bool:
    """
    Skip after resolving a shortcut to its real target workbook.
    """
    name = path.name.lower()
    stem = path.stem.lower()

    if path.suffix.lower() not in EXCEL_EXTS:
        return True

    if name.startswith("~$"):
        return True

    if stem.startswith("current_month_expense_fill_report"):
        return True

    if re.search(r"_backup_\d{8}_\d{6}$", stem):
        return True

    if same_file(path, source_path):
        return True

    return False


def collect_target_workbooks(
    target_input: Path,
    source_path: Path,
    scan_subfolders: bool = False,
) -> List[TargetWorkbook]:
    """
    Return real target workbook files to update.

    Supports:
    - one actual Excel workbook
    - one .lnk shortcut to an Excel workbook
    - one folder containing Excel workbooks
    - one folder containing .lnk shortcuts to Excel workbooks
    """
    if target_input.is_file():
        candidates = [target_input]

    elif target_input.is_dir():
        iterator = target_input.rglob("*") if scan_subfolders else target_input.glob("*")

        candidates = [
            p for p in iterator
            if p.is_file() and not should_skip_target_candidate(p)
        ]

    else:
        raise FileNotFoundError(f"Target workbook/folder not found: {target_input}")

    resolved_items: List[TargetWorkbook] = []
    seen = set()

    for candidate in candidates:
        try:
            resolved = resolve_windows_shortcut(candidate)
        except Exception as exc:
            print(f"[WARN] Skipped shortcut because it could not be resolved: {candidate} | {exc}")
            continue

        if not resolved.exists():
            print(f"[WARN] Skipped because resolved target does not exist: {candidate} -> {resolved}")
            continue

        if should_skip_resolved_workbook(resolved, source_path):
            continue

        try:
            key = str(resolved.resolve()).lower()
        except Exception:
            key = os.path.normcase(os.path.abspath(str(resolved)))

        if key in seen:
            continue

        seen.add(key)

        shortcut_path = candidate if candidate.suffix.lower() == SHORTCUT_EXT else None

        resolved_items.append(
            TargetWorkbook(
                actual_path=resolved,
                shortcut_path=shortcut_path,
            )
        )

        if shortcut_path:
            print(f"[STEP] Resolved shortcut: {candidate.name} -> {resolved}")

    return sorted(resolved_items, key=lambda item: str(item.actual_path).lower())


def choose_output_dir(target_input: Path, output_dir_arg: Optional[str]) -> Path:
    if output_dir_arg:
        return Path(strip_quotes(output_dir_arg)).expanduser()

    if target_input.is_dir():
        return target_input

    return target_input.parent


def best_column_by_nonempty(rows: Sequence[Sequence[object]], candidate_indices: List[int]) -> Optional[int]:
    if not candidate_indices:
        return None

    best_idx = candidate_indices[0]
    best_count = -1

    # Inspect up to 500 rows after the header.
    for idx in candidate_indices:
        count = 0

        for row in rows[1:501]:
            if idx < len(row) and str(row[idx] or "").strip():
                count += 1

        if count > best_count:
            best_count = count
            best_idx = idx

    return best_idx


def available_source_month_keys(headers: Sequence[str]) -> List[str]:
    """
    Return month keys found in source headers like APR-2026 Actual Expenditure Amount.
    """
    out: List[str] = []
    seen = set()

    month_re = re.compile(
        r"\b(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|SEPT|OCT|NOV|DEC)[A-Z]*[-\s/]*(20\d{2}|\d{2})\b"
    )

    for h in headers:
        if "ACTUAL EXPENDITURE AMOUNT" not in h:
            continue

        m = month_re.search(h)

        if not m:
            continue

        mon = m.group(1)

        if mon == "SEPT":
            mon = "SEP"

        year = m.group(2)

        if len(year) == 2:
            year = "20" + year

        key = f"{mon}-{year}"

        if key not in seen:
            seen.add(key)
            out.append(key)

    return out


def find_source_columns(rows: Sequence[Sequence[object]], month_key: str) -> Tuple[int, Dict[str, int]]:
    """
    Find the source header row and required columns.

    Returns:
        zero-based header row index,
        zero-based column index mapping.
    """
    best_debug = None

    for header_idx, row in enumerate(rows[:50]):
        headers = [normalize_header(v) for v in row]

        if not any(headers):
            continue

        award_candidates = [i for i, h in enumerate(headers) if h == "AWARD NUMBER"]
        task_candidates = [i for i, h in enumerate(headers) if h == "TASK NUMBER"]

        ftd_candidates = [
            i for i, h in enumerate(headers)
            if h in {"FTD OR PTD ACTUAL EXPENDITURE AMOUNT", "PTD ACTUAL EXPENDITURE AMOUNT"}
        ]

        month_actual_candidates = []

        for i, h in enumerate(headers):
            if month_key in h and "ACTUAL EXPENDITURE AMOUNT" in h:
                month_actual_candidates.append(i)

        project_number_candidates = [i for i, h in enumerate(headers) if h == "PROJECT NUMBER"]
        project_owner_candidates = [i for i, h in enumerate(headers) if h == "PROJECT OWNER FULL NAME"]
        task_name_candidates = [i for i, h in enumerate(headers) if h == "TASK NAME"]
        award_name_candidates = [i for i, h in enumerate(headers) if h in {"AWARD NAME", "AWARD LONG NAME"}]

        found_count = sum(
            bool(x)
            for x in [award_candidates, task_candidates, ftd_candidates, month_actual_candidates]
        )

        available_months = available_source_month_keys(headers)

        if best_debug is None or found_count > best_debug[0]:
            best_debug = (
                found_count,
                header_idx + 1,
                [h for h in headers if h][:80],
                available_months,
            )

        if award_candidates and task_candidates and ftd_candidates and month_actual_candidates:
            data_rows_for_scoring = rows[header_idx:]

            award_col = best_column_by_nonempty(data_rows_for_scoring, award_candidates)
            task_col = best_column_by_nonempty(data_rows_for_scoring, task_candidates)

            if award_col is None or task_col is None:
                continue

            cols: Dict[str, int] = {
                "award": award_col,
                "task": task_col,
                "month_actual": month_actual_candidates[0],
                "ftd_ptd_actual": ftd_candidates[0],
            }

            if project_number_candidates:
                cols["project_number"] = project_number_candidates[0]

            if project_owner_candidates:
                cols["project_owner_full_name"] = project_owner_candidates[0]

            if task_name_candidates:
                cols["task_name"] = task_name_candidates[0]

            if award_name_candidates:
                cols["award_name"] = award_name_candidates[0]

            return header_idx, cols

    if best_debug:
        available = best_debug[3] if len(best_debug) > 3 else []

        available_msg = (
            f" Available monthly Actual Expenditure columns found: {', '.join(available)}."
            if available
            else " No monthly Actual Expenditure columns were detected."
        )

        raise KeyError(
            "Could not find all required source columns. "
            f"The source workbook does not contain the requested monthly Actual Expenditure column: {month_key}."
            f"{available_msg} "
            f"Best header guess was row {best_debug[1]} with {best_debug[0]} of 4 required groups. "
            f"Headers seen: {best_debug[2]}"
        )

    raise KeyError("Could not find a usable header row in the source workbook.")


def rows_from_openpyxl_workbook(path: Path) -> List[Tuple[str, List[List[object]]]]:
    wb = load_workbook(path, data_only=True, read_only=True)

    output: List[Tuple[str, List[List[object]]]] = []

    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        output.append((ws.title, rows))

    wb.close()

    return output


def excel_com_available() -> bool:
    if os.name != "nt":
        return False

    try:
        import win32com.client  # type: ignore
        return True
    except Exception:
        return False


def get_excel_app():
    if os.name != "nt":
        raise RuntimeError("Microsoft Excel COM is only available on Windows.")

    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except Exception as exc:
        raise RuntimeError(
            "Microsoft Excel + pywin32 are required for .xls files. "
            "Run: python -m pip install pywin32"
        ) from exc

    pythoncom.CoInitialize()

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    return excel, pythoncom


def rows_from_excel_com_workbook(path: Path) -> List[Tuple[str, List[List[object]]]]:
    excel, pythoncom = get_excel_app()
    wb = None

    try:
        wb = excel.Workbooks.Open(str(path), UpdateLinks=False, ReadOnly=True)

        output: List[Tuple[str, List[List[object]]]] = []

        for ws in wb.Worksheets:
            used = ws.UsedRange
            values = used.Value

            if values is None:
                rows: List[List[object]] = []
            elif not isinstance(values, tuple):
                rows = [[values]]
            else:
                rows = [
                    list(row) if isinstance(row, tuple) else [row]
                    for row in values
                ]

            output.append((str(ws.Name), rows))

        return output

    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)

        excel.Quit()
        pythoncom.CoUninitialize()


def read_source_data(source_path: Path, month_key: str) -> Dict[Tuple[str, str], SourcePTAData]:
    log(f"[STEP] Reading source workbook: {source_path}")

    suffix = source_path.suffix.lower()

    if suffix in OPENPYXL_EXTS:
        sheet_rows = rows_from_openpyxl_workbook(source_path)
    elif suffix == ".xls":
        sheet_rows = rows_from_excel_com_workbook(source_path)
    else:
        raise ValueError(f"Unsupported source workbook extension: {source_path.suffix}")

    best_error: Optional[Exception] = None

    for sheet_name, rows in sheet_rows:
        if not rows:
            continue

        try:
            header_idx, cols = find_source_columns(rows, month_key)
        except Exception as exc:
            best_error = exc
            continue

        log(f"[STEP] Using source sheet '{sheet_name}' with header row {header_idx + 1}")

        index: Dict[Tuple[str, str], SourcePTAData] = {}

        for row in rows[header_idx + 1:]:
            award_col = cols["award"]
            task_col = cols["task"]

            award = normalize_award(row[award_col] if award_col < len(row) else None)
            task = normalize_task(row[task_col] if task_col < len(row) else None)

            if not award or not task:
                continue

            key = (award, task)

            if key not in index:
                index[key] = SourcePTAData(award=award, task=task)

            item = index[key]

            item.month_actual = round(
                item.month_actual + safe_float(
                    row[cols["month_actual"]] if cols["month_actual"] < len(row) else None
                ),
                2,
            )

            item.ftd_ptd_actual = round(
                item.ftd_ptd_actual + safe_float(
                    row[cols["ftd_ptd_actual"]] if cols["ftd_ptd_actual"] < len(row) else None
                ),
                2,
            )

            item.row_count += 1

            for optional_key in ["project_number", "project_owner_full_name", "task_name", "award_name"]:
                col = cols.get(optional_key)

                if col is not None and col < len(row):
                    val = str(row[col] or "").strip()

                    if val and not getattr(item, optional_key):
                        setattr(item, optional_key, val)

        if index:
            log(f"[STEP] Parsed {len(index)} aggregated award/task pair(s) from source.")
            return index

        best_error = ValueError(f"Source sheet '{sheet_name}' had headers but no parsable award/task rows.")

    if best_error:
        raise best_error

    raise ValueError("No usable source sheet was found.")


def zero_source_data(award: str, task: str) -> SourcePTAData:
    return SourcePTAData(
        award=award,
        task=task,
        month_actual=0.0,
        ftd_ptd_actual=0.0,
        row_count=0,
    )


def find_current_expense_row_openpyxl(ws) -> Optional[int]:
    for row_idx in range(1, ws.max_row + 1):
        if current_month_expense_label(ws.cell(row=row_idx, column=1).value):
            return row_idx

    return None


def find_expense_row_openpyxl(ws) -> Optional[int]:
    for row_idx in range(1, ws.max_row + 1):
        if expense_label(ws.cell(row=row_idx, column=1).value):
            return row_idx

    return None


def find_month_column_openpyxl(
    ws,
    year: int,
    month: int,
    current_row: int,
) -> Tuple[Optional[int], Optional[int]]:
    """
    Return:
        header_row,
        month_col

    Prioritizes rows above/near the target row.
    """
    max_scan_row = min(ws.max_row, max(current_row, 50))

    candidates: List[Tuple[int, int, int]] = []

    for r in range(1, max_scan_row + 1):
        for c in range(1, ws.max_column + 1):
            if month_matches(ws.cell(row=r, column=c).value, year, month):
                score = abs(current_row - r)

                # Prefer header rows above current row.
                if r <= current_row:
                    score -= 10000

                candidates.append((score, r, c))

    if not candidates:
        return None, None

    candidates.sort(key=lambda x: (x[0], x[1], x[2]))

    _, header_row, month_col = candidates[0]

    return header_row, month_col


def copy_style_from_prior_month_openpyxl(
    ws,
    header_row: int,
    current_row: int,
    month_col: int,
) -> str:
    """
    Copy formatting from the nearest prior month cell in the same row, if found.
    """
    target = ws.cell(row=current_row, column=month_col)

    for c in range(month_col - 1, 0, -1):
        if is_any_month_header(ws.cell(row=header_row, column=c).value):
            src = ws.cell(row=current_row, column=c)

            if src.has_style:
                target._style = copy(src._style)

            if src.number_format:
                target.number_format = src.number_format

            if src.alignment:
                target.alignment = copy(src.alignment)

            return f"{get_column_letter(c)}{current_row}"

    return ""


def process_target_openpyxl(
    target_path: Path,
    source_index: Dict[Tuple[str, str], SourcePTAData],
    year: int,
    month: int,
    shortcut_path: Optional[Path] = None,
) -> List[SheetResult]:
    keep_vba = target_path.suffix.lower() == ".xlsm"

    wb = load_workbook(target_path, keep_vba=keep_vba)

    results: List[SheetResult] = []
    changed = False

    try:
        for ws in wb.worksheets:
            award = normalize_award(ws["B3"].value)
            task = normalize_task(ws["B5"].value)

            base = SheetResult(
                status="SKIP",
                target_workbook=str(target_path),
                shortcut_path=str(shortcut_path or ""),
                sheet=ws.title,
                award=award,
                task=task,
            )

            if not award or not task:
                base.notes = "Skipped: B3 award or B5 task is blank/not usable."
                results.append(base)
                continue

            current_row = find_current_expense_row_openpyxl(ws)

            if current_row is None:
                base.notes = "Skipped: column A does not contain CURRENT MONTH EXPENSE / CURRENT MO. EXPENSE."
                results.append(base)
                continue

            header_row, month_col = find_month_column_openpyxl(ws, year, month, current_row)

            if not header_row or not month_col:
                base.notes = "Skipped: requested month column was not found on this worksheet."
                results.append(base)
                continue

            pdata = source_index.get((award, task), zero_source_data(award, task))

            target_cell = ws.cell(row=current_row, column=month_col)
            previous_value = target_cell.value

            style_source = copy_style_from_prior_month_openpyxl(ws, header_row, current_row, month_col)

            target_cell.value = float(pdata.month_actual)
            changed = True

            # Validation:
            # CURRENT MONTH EXPENSE should equal the source monthly actual column.
            # Workbook cumulative EXPENSE row should equal source FTD/PTD actual.
            current_month_diff = round(safe_float(target_cell.value) - pdata.month_actual, 2)

            expense_row = find_expense_row_openpyxl(ws)
            workbook_expense_value: Optional[float] = None
            expense_diff: Optional[float] = None
            expense_cell = ""

            if expense_row is not None:
                workbook_expense_value = safe_float(ws.cell(row=expense_row, column=month_col).value)
                expense_diff = round(workbook_expense_value - pdata.ftd_ptd_actual, 2)
                expense_cell = f"{get_column_letter(month_col)}{expense_row}"

            if pdata.row_count == 0:
                status = "OK"
                notes = "No matching source rows were found for this award/task, so 0 was written."

            elif abs(current_month_diff) > TOLERANCE:
                status = "CHECK"
                notes = (
                    f"CURRENT MONTH EXPENSE was expected to equal source monthly actual "
                    f"{pdata.month_actual:,.2f}, but workbook shows {safe_float(target_cell.value):,.2f}. "
                    "Value was kept; please review."
                )

            elif expense_row is None:
                status = "CHECK"
                notes = (
                    "CURRENT MONTH EXPENSE was filled from the monthly Actual Expenditure Amount column, "
                    "but the cumulative EXPENSE row was not found, so FTD/PTD validation could not be completed."
                )

            elif expense_diff is not None and abs(expense_diff) <= TOLERANCE:
                status = "OK"
                notes = (
                    "CURRENT MONTH EXPENSE was filled from the monthly Actual Expenditure Amount column. "
                    "Workbook EXPENSE row matches source FTD/PTD Actual Expenditure Amount."
                )

            else:
                status = "CHECK"
                notes = (
                    "CURRENT MONTH EXPENSE was filled from the monthly Actual Expenditure Amount column, "
                    f"but workbook EXPENSE row {format_money(workbook_expense_value)} does not match "
                    f"source FTD/PTD Actual Expenditure Amount {pdata.ftd_ptd_actual:,.2f}. "
                    "Value was kept; please review."
                )

            if style_source:
                notes += f" Format copied from {style_source}."

            results.append(
                SheetResult(
                    status=status,
                    target_workbook=str(target_path),
                    shortcut_path=str(shortcut_path or ""),
                    sheet=ws.title,
                    award=award,
                    task=task,
                    source_rows=pdata.row_count,
                    source_month_actual=pdata.month_actual,
                    source_ftd_ptd_actual=pdata.ftd_ptd_actual,
                    written_value=pdata.month_actual,
                    previous_value=previous_value,
                    workbook_expense_value=workbook_expense_value,
                    difference_current_month_vs_source_month=current_month_diff,
                    difference_expense_vs_ftd_ptd=expense_diff,
                    month_cell=f"{get_column_letter(month_col)}{current_row}",
                    expense_cell=expense_cell,
                    month_header_cell=f"{get_column_letter(month_col)}{header_row}",
                    project_number=pdata.project_number,
                    project_owner_full_name=pdata.project_owner_full_name,
                    task_name=pdata.task_name,
                    award_name=pdata.award_name,
                    notes=notes,
                )
            )

        if changed:
            wb.save(target_path)

        return results

    finally:
        wb.close()


def excel_col(n: int) -> str:
    out = ""

    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out

    return out


def find_current_expense_row_com(ws, max_row: int) -> Optional[int]:
    for r in range(1, max_row + 1):
        if current_month_expense_label(ws.Cells(r, 1).Value):
            return r

    return None


def find_expense_row_com(ws, max_row: int) -> Optional[int]:
    for r in range(1, max_row + 1):
        if expense_label(ws.Cells(r, 1).Value):
            return r

    return None


def find_month_column_com(
    ws,
    year: int,
    month: int,
    current_row: int,
    max_row: int,
    max_col: int,
) -> Tuple[Optional[int], Optional[int]]:
    max_scan_row = min(max_row, max(current_row, 50))

    candidates: List[Tuple[int, int, int]] = []

    for r in range(1, max_scan_row + 1):
        for c in range(1, max_col + 1):
            if month_matches(ws.Cells(r, c).Value, year, month):
                score = abs(current_row - r)

                if r <= current_row:
                    score -= 10000

                candidates.append((score, r, c))

    if not candidates:
        return None, None

    candidates.sort(key=lambda x: (x[0], x[1], x[2]))

    _, header_row, month_col = candidates[0]

    return header_row, month_col


def copy_style_from_prior_month_com(
    excel,
    ws,
    header_row: int,
    current_row: int,
    month_col: int,
) -> str:
    target = ws.Cells(current_row, month_col)

    for c in range(month_col - 1, 0, -1):
        if is_any_month_header(ws.Cells(header_row, c).Value):
            ws.Cells(current_row, c).Copy()
            target.PasteSpecial(Paste=-4122)  # xlPasteFormats
            excel.CutCopyMode = False

            return f"{excel_col(c)}{current_row}"

    return ""


def process_target_excel_com(
    target_path: Path,
    source_index: Dict[Tuple[str, str], SourcePTAData],
    year: int,
    month: int,
    shortcut_path: Optional[Path] = None,
) -> List[SheetResult]:
    excel, pythoncom = get_excel_app()

    wb = None
    results: List[SheetResult] = []
    changed = False

    try:
        wb = excel.Workbooks.Open(str(target_path), UpdateLinks=False, ReadOnly=False)

        for ws in wb.Worksheets:
            sheet_name = str(ws.Name)

            award = normalize_award(ws.Range("B3").Value)
            task = normalize_task(ws.Range("B5").Value)

            base = SheetResult(
                status="SKIP",
                target_workbook=str(target_path),
                shortcut_path=str(shortcut_path or ""),
                sheet=sheet_name,
                award=award,
                task=task,
            )

            if not award or not task:
                base.notes = "Skipped: B3 award or B5 task is blank/not usable."
                results.append(base)
                continue

            used = ws.UsedRange
            max_row = used.Row + used.Rows.Count - 1
            max_col = used.Column + used.Columns.Count - 1

            current_row = find_current_expense_row_com(ws, max_row)

            if current_row is None:
                base.notes = "Skipped: column A does not contain CURRENT MONTH EXPENSE / CURRENT MO. EXPENSE."
                results.append(base)
                continue

            header_row, month_col = find_month_column_com(ws, year, month, current_row, max_row, max_col)

            if not header_row or not month_col:
                base.notes = "Skipped: requested month column was not found on this worksheet."
                results.append(base)
                continue

            pdata = source_index.get((award, task), zero_source_data(award, task))

            target_rng = ws.Cells(current_row, month_col)
            previous_value = target_rng.Value

            style_source = copy_style_from_prior_month_com(excel, ws, header_row, current_row, month_col)

            target_rng.Value = float(pdata.month_actual)
            changed = True

            try:
                excel.Calculate()
            except Exception:
                pass

            current_month_diff = round(safe_float(target_rng.Value) - pdata.month_actual, 2)

            expense_row = find_expense_row_com(ws, max_row)
            workbook_expense_value: Optional[float] = None
            expense_diff: Optional[float] = None
            expense_cell = ""

            if expense_row is not None:
                workbook_expense_value = safe_float(ws.Cells(expense_row, month_col).Value)
                expense_diff = round(workbook_expense_value - pdata.ftd_ptd_actual, 2)
                expense_cell = f"{excel_col(month_col)}{expense_row}"

            if pdata.row_count == 0:
                status = "OK"
                notes = "No matching source rows were found for this award/task, so 0 was written."

            elif abs(current_month_diff) > TOLERANCE:
                status = "CHECK"
                notes = (
                    f"CURRENT MONTH EXPENSE was expected to equal source monthly actual "
                    f"{pdata.month_actual:,.2f}, but workbook shows {safe_float(target_rng.Value):,.2f}. "
                    "Value was kept; please review."
                )

            elif expense_row is None:
                status = "CHECK"
                notes = (
                    "CURRENT MONTH EXPENSE was filled from the monthly Actual Expenditure Amount column, "
                    "but the cumulative EXPENSE row was not found, so FTD/PTD validation could not be completed."
                )

            elif expense_diff is not None and abs(expense_diff) <= TOLERANCE:
                status = "OK"
                notes = (
                    "CURRENT MONTH EXPENSE was filled from the monthly Actual Expenditure Amount column. "
                    "Workbook EXPENSE row matches source FTD/PTD Actual Expenditure Amount."
                )

            else:
                status = "CHECK"
                notes = (
                    "CURRENT MONTH EXPENSE was filled from the monthly Actual Expenditure Amount column, "
                    f"but workbook EXPENSE row {format_money(workbook_expense_value)} does not match "
                    f"source FTD/PTD Actual Expenditure Amount {pdata.ftd_ptd_actual:,.2f}. "
                    "Value was kept; please review."
                )

            if style_source:
                notes += f" Format copied from {style_source}."

            results.append(
                SheetResult(
                    status=status,
                    target_workbook=str(target_path),
                    shortcut_path=str(shortcut_path or ""),
                    sheet=sheet_name,
                    award=award,
                    task=task,
                    source_rows=pdata.row_count,
                    source_month_actual=pdata.month_actual,
                    source_ftd_ptd_actual=pdata.ftd_ptd_actual,
                    written_value=pdata.month_actual,
                    previous_value=previous_value,
                    workbook_expense_value=workbook_expense_value,
                    difference_current_month_vs_source_month=current_month_diff,
                    difference_expense_vs_ftd_ptd=expense_diff,
                    month_cell=f"{excel_col(month_col)}{current_row}",
                    expense_cell=expense_cell,
                    month_header_cell=f"{excel_col(month_col)}{header_row}",
                    project_number=pdata.project_number,
                    project_owner_full_name=pdata.project_owner_full_name,
                    task_name=pdata.task_name,
                    award_name=pdata.award_name,
                    notes=notes,
                )
            )

        if changed:
            wb.Save()

        return results

    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)

        excel.Quit()
        pythoncom.CoUninitialize()


def write_processing_report(output_dir: Path, results: List[SheetResult], month_key: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = output_dir / f"current_month_expense_fill_report_{timestamp}.xlsx"

    headers = [
        "status",
        "target_workbook",
        "shortcut_path",
        "sheet",
        "award",
        "task",
        "source_rows",
        f"source_{month_key}_actual_expenditure_amount",
        "source_ftd_or_ptd_actual_expenditure_amount",
        "written_current_month_expense",
        "previous_current_month_expense",
        "workbook_expense_value",
        "difference_current_month_vs_source_month",
        "difference_expense_vs_ftd_ptd",
        "current_month_expense_cell",
        "expense_cell",
        "month_header_cell",
        "project_number",
        "project_owner_full_name",
        "task_name",
        "award_name",
        "notes",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Processing Report"

    ws.append(headers)

    for r in results:
        ws.append(
            [
                r.status,
                r.target_workbook,
                r.shortcut_path,
                r.sheet,
                r.award,
                r.task,
                r.source_rows,
                r.source_month_actual,
                r.source_ftd_ptd_actual,
                r.written_value,
                r.previous_value,
                r.workbook_expense_value,
                r.difference_current_month_vs_source_month,
                r.difference_expense_vs_ftd_ptd,
                r.month_cell,
                r.expense_cell,
                r.month_header_cell,
                r.project_number,
                r.project_owner_full_name,
                r.task_name,
                r.award_name,
                r.notes,
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    status_col = headers.index("status") + 1

    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row_idx, status_col).value or "").upper()

        if status == "OK":
            fill = PatternFill("solid", fgColor="E2F0D9")
        elif status == "CHECK":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif status == "SKIP":
            fill = PatternFill("solid", fgColor="E7E6E6")
        else:
            fill = PatternFill("solid", fgColor="FCE4D6")

        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).fill = fill

    currency_headers = {
        f"source_{month_key}_actual_expenditure_amount",
        "source_ftd_or_ptd_actual_expenditure_amount",
        "written_current_month_expense",
        "previous_current_month_expense",
        "workbook_expense_value",
        "difference_current_month_vs_source_month",
        "difference_expense_vs_ftd_ptd",
    }

    for col_idx, header in enumerate(headers, 1):
        if header in currency_headers:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = '#,##0.00;[Red](#,##0.00);-'

        max_len = len(header)

        for row_idx in range(2, min(ws.max_row, 200) + 1):
            max_len = max(max_len, len(str(ws.cell(row_idx, col_idx).value or "")))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 70)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(report_path)

    return report_path


def ask_path(prompt: str) -> str:
    return strip_quotes(input(prompt).strip())


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Fill CURRENT MONTH EXPENSE by worksheet/tab from Expenditure Summary raw data."
    )

    parser.add_argument(
        "--target",
        help="Path to the target workbook, .lnk shortcut, or folder containing target workbooks/shortcuts.",
    )
    parser.add_argument(
        "--source",
        help="Path to raw Expenditure Summary workbook.",
    )
    parser.add_argument(
        "--month",
        help="Target workbook month column to fill, e.g. may-2026.",
    )
    parser.add_argument(
        "--source-month",
        help="Source Actual Expenditure month to read, e.g. apr-2026. Defaults to --month.",
    )
    parser.add_argument(
        "--output-dir",
        help="Folder where the processing report should be written. Defaults to target workbook/folder.",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Accepted for compatibility only. Backups are always disabled.",
    )
    parser.add_argument(
        "--use-excel-com",
        action="store_true",
        help="Use Microsoft Excel COM for all target workbooks, useful for .xls or formulas/macros.",
    )
    parser.add_argument(
        "--scan-subfolders",
        action="store_true",
        help="If target is a folder, also scan subfolders.",
    )

    args = parser.parse_args()

    target_was_prompted = False

    if not args.target:
        args.target = ask_path("Enter FULL path to the target workbook, shortcut, OR target folder to update: ")
        target_was_prompted = True

    if not args.source:
        args.source = ask_path("Enter FULL path to the raw Expenditure Summary workbook: ")

    if not args.month:
        args.month = strip_quotes(input("Enter target workbook month column to fill, e.g. may-2026: "))

    if not args.source_month:
        args.source_month = strip_quotes(
            input("Enter source Actual Expenditure month to read, or press Enter to use the same month: ")
        )

        if not args.source_month:
            args.source_month = args.month

    target_input = Path(strip_quotes(args.target)).expanduser()
    source_path = Path(strip_quotes(args.source)).expanduser()

    if target_was_prompted and target_input.is_dir() and not args.scan_subfolders:
        scan_answer = strip_quotes(input("Scan subfolders too? Type Y for yes, or press Enter for no: "))
        args.scan_subfolders = scan_answer.strip().lower() == "y"

    output_dir = choose_output_dir(target_input, args.output_dir)

    print("\n=== RUNNING WITH ===")
    print(f"Target input:    {target_input}")
    print(f"Source workbook: {source_path}")
    print(f"Target month:    {args.month}")
    print(f"Source month:    {args.source_month or args.month}")
    print(f"Report folder:   {output_dir}")
    print(f"Scan subfolders: {'Yes' if args.scan_subfolders else 'No'}")
    print("====================\n")

    if not source_path.exists():
        print(f"[FAIL] Source workbook not found: {source_path}")
        return 1

    if source_path.suffix.lower() not in EXCEL_EXTS:
        print(f"[FAIL] Unsupported source workbook type: {source_path.suffix}")
        return 1

    try:
        target_workbooks = collect_target_workbooks(
            target_input=target_input,
            source_path=source_path,
            scan_subfolders=args.scan_subfolders,
        )
    except Exception as exc:
        print(f"[FAIL] Could not collect target workbooks: {exc}")
        return 1

    if not target_workbooks:
        print("[FAIL] No target Excel workbooks or usable shortcuts were found.")
        print("[HINT] If your folder shows Type = Shortcut, this version should resolve .lnk files.")
        print("[HINT] If it still finds none, right-click one shortcut > Properties and check the Target path.")
        return 1

    print(f"[STEP] Found {len(target_workbooks)} target workbook(s):")

    for item in target_workbooks:
        if item.shortcut_path:
            print(f"  - {item.actual_path}  [from shortcut: {item.shortcut_path.name}]")
        else:
            print(f"  - {item.actual_path}")

    try:
        year, month, target_month_key, display_month = parse_month_input(args.month)
    except Exception as exc:
        print(f"[FAIL] Target month could not be parsed: {exc}")
        return 1

    source_month_text = args.source_month or args.month

    try:
        _source_year, _source_month, source_month_key, _source_display_month = parse_month_input(source_month_text)
    except Exception as exc:
        print(f"[FAIL] Source month could not be parsed: {exc}")
        return 1

    try:
        source_index = read_source_data(source_path, source_month_key)
    except Exception as exc:
        print(f"[FAIL] Source workbook could not be read for source month {source_month_key}: {exc}")
        return 1

    needs_com = args.use_excel_com or any(item.actual_path.suffix.lower() == ".xls" for item in target_workbooks)

    if needs_com and not excel_com_available():
        print("[FAIL] Excel COM was requested or required for .xls workbooks, but Microsoft Excel + pywin32 is not available.")
        print("Install pywin32 with: python -m pip install pywin32")
        return 1

    all_results: List[SheetResult] = []
    backup_paths: List[Path] = []

    for item in target_workbooks:
        target_path = item.actual_path
        shortcut_path = item.shortcut_path

        print(f"\n[STEP] Processing target workbook: {target_path}")

        if shortcut_path:
            print(f"[STEP] Source shortcut: {shortcut_path}")

        if not target_path.exists():
            all_results.append(
                SheetResult(
                    status="FAIL",
                    target_workbook=str(target_path),
                    shortcut_path=str(shortcut_path or ""),
                    sheet="",
                    notes="Target workbook was not found at processing time.",
                )
            )
            continue

        # Backups are intentionally disabled.
        # This program edits the target workbook directly and does not create backup files.
        print("[STEP] Backup creation is disabled. No backup file will be created.")

        try:
            if target_path.suffix.lower() == ".xls" or args.use_excel_com:
                results = process_target_excel_com(
                    target_path=target_path,
                    source_index=source_index,
                    year=year,
                    month=month,
                    shortcut_path=shortcut_path,
                )
            else:
                results = process_target_openpyxl(
                    target_path=target_path,
                    source_index=source_index,
                    year=year,
                    month=month,
                    shortcut_path=shortcut_path,
                )

            all_results.extend(results)

            updated_in_file = sum(1 for r in results if r.written_value is not None)
            checks_in_file = sum(1 for r in results if r.status == "CHECK")
            skips_in_file = sum(1 for r in results if r.status == "SKIP")

            print(
                f"[STEP] Finished {target_path.name}: "
                f"updated {updated_in_file}, check {checks_in_file}, skip {skips_in_file}"
            )

        except Exception as exc:
            print(f"[FAIL] Target workbook could not be processed: {target_path} | {exc}")

            all_results.append(
                SheetResult(
                    status="FAIL",
                    target_workbook=str(target_path),
                    shortcut_path=str(shortcut_path or ""),
                    sheet="",
                    notes=f"Target workbook could not be processed. Reason: {exc}",
                )
            )

    try:
        report_path = write_processing_report(output_dir, all_results, source_month_key)
        print(f"\n[STEP] Processing report written to: {report_path}")
    except Exception as exc:
        print(f"[FAIL] Could not write processing report: {exc}")
        return 1

    updated = sum(1 for r in all_results if r.written_value is not None)
    ok_count = sum(1 for r in all_results if r.status == "OK")
    check_count = sum(1 for r in all_results if r.status == "CHECK")
    skip_count = sum(1 for r in all_results if r.status == "SKIP")
    fail_count = sum(1 for r in all_results if r.status == "FAIL")

    print("\n=== SUMMARY ===")
    print(f"Target workbooks found: {len(target_workbooks)}")
    print("Backups created:        0 -- backup creation is disabled")
    print(f"Sheets updated:         {updated}")
    print(f"OK:                     {ok_count}")
    print(f"CHECK:                  {check_count}")
    print(f"SKIP:                   {skip_count}")
    print(f"FAIL:                   {fail_count}")

    if check_count:
        print(
            "\nItems needing review because workbook EXPENSE row did not match "
            "FTD/PTD Actual Expenditure Amount, or validation row was missing:"
        )

        for r in all_results:
            if r.status == "CHECK":
                print(
                    f"  - {Path(r.target_workbook).name} | {r.sheet}: {r.award}/{r.task} | "
                    f"current month written {format_money(r.written_value)} | "
                    f"workbook EXPENSE {format_money(r.workbook_expense_value)} "
                    f"vs source FTD/PTD {r.source_ftd_ptd_actual:,.2f}"
                )

    if fail_count:
        return 2

    if check_count:
        return 3

    return 0


if __name__ == "__main__":
    raise SystemExit(main())