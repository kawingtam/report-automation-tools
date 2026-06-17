#!/usr/bin/env python3
"""
Fill CURRENT MONTH COMMIT for multi-PTA balance workbooks from a Commt PO Report export.

What this program does:
- Accepts a target workbook, a target folder, or one/more .lnk shortcuts.
- If a folder is entered, scans Excel workbooks and .lnk shortcuts inside the folder.
- Resolves Windows .lnk shortcuts and edits the shortcut target workbook.
- Does NOT create backup files.
- For each worksheet/tab in each target workbook, reads:
    B3 = Award
    B5 = Task
- Skips tabs where B3/B5 are not usable or where CURRENT MONTH COMMIT / month column is not found.
- Finds the row in column A labeled CURRENT MONTH COMMIT / CURRENT MO. COMMIT.
- Finds the month column from the entered month, such as apr-2026 or APR-2026.
- Aggregates the Commt PO Report source by Award Number + Task Number.
- Writes the aggregated Expenditure Commitment Amount into the CURRENT MONTH COMMIT/month cell.
- If no matching source rows are found for a worksheet Award/Task, writes 0.
- Writes a processing report.

Requirements:
- Python packages: openpyxl
- For .xls target workbooks or .lnk shortcuts, run on Windows with Microsoft Excel and pywin32 installed.
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import subprocess
import sys
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TOLERANCE = 0.01
EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}
OPENPYXL_EXTS = {".xlsx", ".xlsm"}
INPUT_EXTS = EXCEL_EXTS | {".lnk"}


@dataclass
class CommitSourceData:
    award: str
    task: str
    commitment_amount: float = 0.0
    row_count: int = 0
    project_number: str = ""
    project_owner_full_name: str = ""
    task_name: str = ""
    award_name: str = ""
    po_numbers: str = ""


@dataclass
class WorkItem:
    input_path: Path
    workbook_path: Path
    source_type: str = "file"


@dataclass
class SheetResult:
    status: str
    input_path: str
    target_workbook: str
    sheet: str
    award: str = ""
    task: str = ""
    source_rows: int = 0
    source_commitment_amount: float = 0.0
    written_value: Optional[float] = None
    previous_value: object = ""
    difference_written_vs_source: Optional[float] = None
    current_month_commit_cell: str = ""
    month_header_cell: str = ""
    style_source_cell: str = ""
    project_number: str = ""
    project_owner_full_name: str = ""
    task_name: str = ""
    award_name: str = ""
    po_numbers: str = ""
    notes: str = ""


def log(msg: str) -> None:
    print(msg)


def strip_quotes(value) -> str:
    return str(value or "").strip().strip('"').strip("'")


def clean_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_header(value) -> str:
    return clean_header(value).upper()


def normalize_label(value) -> str:
    text = str(value or "").strip().upper()
    text = text.replace(".", " ")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_award(value) -> str:
    """Normalize award from B3 or source. Keep full award token; do not assume 5 characters."""
    text = str(value or "").strip().upper()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    tokens = re.findall(r"[A-Z0-9]+", text)
    if not tokens:
        return ""
    skip = {"AWARD", "NUMBER", "NO", "TASK", "PROJECT", "LEGACY", "NA", "NONE"}
    usable = [tok for tok in tokens if tok not in skip]
    return usable[-1] if usable else ""


def normalize_task(value) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    m = re.search(r"\d+", text)
    return m.group(0) if m else ""


def safe_float(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    neg = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    try:
        number = float(text)
        return -number if neg else number
    except Exception:
        return 0.0


def parse_month_input(month_text: str) -> Tuple[int, int, str]:
    cleaned = strip_quotes(month_text).replace("_", "-").replace("/", "-")
    cleaned = re.sub(r"\s+", " ", cleaned)
    formats = [
        "%b-%Y", "%b-%y", "%B-%Y", "%B-%y",
        "%b %Y", "%b %y", "%B %Y", "%B %y",
        "%Y-%m", "%Y %m", "%m-%Y", "%m-%y",
    ]
    for fmt in formats:
        try:
            parsed = dt.datetime.strptime(cleaned, fmt)
            return parsed.year, parsed.month, parsed.strftime("%b-%Y").upper()
        except ValueError:
            continue
    raise ValueError(f"Could not parse month '{month_text}'. Try values like apr-2026 or 2026-04.")


def month_matches(value, year: int, month: int) -> bool:
    if isinstance(value, (dt.datetime, dt.date)):
        return value.year == year and value.month == month
    text = str(value or "").strip()
    if not text:
        return False
    text = text.replace("/", "-")
    text = re.sub(r"\s+", " ", text)
    for fmt in ["%b-%Y", "%b-%y", "%B-%Y", "%B-%y", "%Y-%m", "%m-%Y", "%m-%y"]:
        try:
            parsed = dt.datetime.strptime(text, fmt)
            return parsed.year == year and parsed.month == month
        except ValueError:
            continue
    return False


def is_any_month_header(value) -> bool:
    if isinstance(value, (dt.datetime, dt.date)):
        return True
    text = str(value or "").strip().replace("/", "-")
    if not text:
        return False
    for fmt in ["%b-%Y", "%b-%y", "%B-%Y", "%B-%y", "%Y-%m", "%m-%Y", "%m-%y"]:
        try:
            dt.datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def current_month_commit_label(value) -> bool:
    text = normalize_label(value)
    return text in {
        "CURRENT MONTH COMMIT",
        "CURRENT MONTH COMMITS",
        "CURRENT MO COMMIT",
        "CURRENT MO COMMITS",
        "CURRENT MON COMMIT",
        "CURRENT MON COMMITS",
        "CURRENT MONTH COMMITMENT",
        "CURRENT MONTH COMMITMENTS",
        "CURRENT MO COMMITMENT",
        "CURRENT MO COMMITMENTS",
        "CURRENT MONTH COMMT",
        "CURRENT MO COMMT",
    }


def best_column_by_nonempty(rows: Sequence[Sequence[object]], candidate_indices: List[int]) -> Optional[int]:
    if not candidate_indices:
        return None
    best_idx = candidate_indices[0]
    best_count = -1
    for idx in candidate_indices:
        count = 0
        for row in rows[1:501]:
            if idx < len(row) and str(row[idx] or "").strip():
                count += 1
        if count > best_count:
            best_count = count
            best_idx = idx
    return best_idx


def rows_from_openpyxl_workbook(path: Path) -> List[Tuple[str, List[List[object]]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    output: List[Tuple[str, List[List[object]]]] = []
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        output.append((ws.title, rows))
    return output


def excel_com_available() -> bool:
    if os.name != "nt":
        return False
    try:
        import win32com.client  # type: ignore
        return True
    except Exception:
        return False


def get_excel_app():
    if os.name != "nt":
        raise RuntimeError("Microsoft Excel COM is only available on Windows.")
    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except Exception as exc:
        raise RuntimeError("Microsoft Excel + pywin32 are required for .xls files and .lnk shortcuts. Run: python -m pip install pywin32") from exc
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    return excel, pythoncom


def rows_from_excel_com_workbook(path: Path) -> List[Tuple[str, List[List[object]]]]:
    excel, pythoncom = get_excel_app()
    wb = None
    try:
        wb = excel.Workbooks.Open(str(path), UpdateLinks=False, ReadOnly=True)
        output: List[Tuple[str, List[List[object]]]] = []
        for ws in wb.Worksheets:
            used = ws.UsedRange
            values = used.Value
            if values is None:
                rows: List[List[object]] = []
            elif not isinstance(values, tuple):
                rows = [[values]]
            else:
                rows = [list(row) if isinstance(row, tuple) else [row] for row in values]
            output.append((str(ws.Name), rows))
        return output
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()
        pythoncom.CoUninitialize()


def find_commit_source_columns(rows: Sequence[Sequence[object]]) -> Tuple[int, Dict[str, int]]:
    """Find header row and source columns in the Commt PO Report.

    Required columns:
    - Award Number
    - Task Number
    - Expenditure Commitment Amount
    """
    best_debug = None
    for header_idx, row in enumerate(rows[:50]):
        headers = [normalize_header(v) for v in row]
        if not any(headers):
            continue

        award_candidates = [i for i, h in enumerate(headers) if h == "AWARD NUMBER"]
        task_candidates = [i for i, h in enumerate(headers) if h == "TASK NUMBER"]
        amount_candidates = [
            i for i, h in enumerate(headers)
            if h in {"EXPENDITURE COMMITMENT AMOUNT", "COMMITMENT AMOUNT", "EXPENDITURE COMMIT AMOUNT"}
            or ("COMMITMENT" in h and "AMOUNT" in h and "PERCENT" not in h)
        ]

        found_count = sum(bool(x) for x in [award_candidates, task_candidates, amount_candidates])
        if best_debug is None or found_count > best_debug[0]:
            best_debug = (found_count, header_idx + 1, [h for h in headers if h][:80])

        if award_candidates and task_candidates and amount_candidates:
            data_rows_for_scoring = rows[header_idx:]
            cols: Dict[str, int] = {
                "award": best_column_by_nonempty(data_rows_for_scoring, award_candidates),  # type: ignore[dict-item]
                "task": best_column_by_nonempty(data_rows_for_scoring, task_candidates),  # type: ignore[dict-item]
                "commitment_amount": amount_candidates[0],
            }
            if cols["award"] is None or cols["task"] is None:
                continue

            optional_exact = {
                "project_number": "PROJECT NUMBER",
                "project_owner_full_name": "PROJECT OWNER FULL NAME",
                "task_name": "TASK NAME",
                "award_name": "AWARD NAME",
                "po_number": "PURCHASE ORDER NUMBER",
            }
            for key, header_name in optional_exact.items():
                matches = [i for i, h in enumerate(headers) if h == header_name]
                if matches:
                    cols[key] = matches[0]
            return header_idx, cols

    if best_debug:
        raise KeyError(
            "Could not find all required Commt PO Report columns. "
            f"Best header guess was row {best_debug[1]} with {best_debug[0]} of 3 required groups. "
            f"Headers seen: {best_debug[2]}"
        )
    raise KeyError("Could not find a usable header row in the Commt PO Report source workbook.")


def read_commit_source_data(source_path: Path) -> Dict[Tuple[str, str], CommitSourceData]:
    log(f"[STEP] Reading Commt PO Report source workbook: {source_path}")
    suffix = source_path.suffix.lower()
    if suffix in OPENPYXL_EXTS:
        sheet_rows = rows_from_openpyxl_workbook(source_path)
    elif suffix == ".xls":
        sheet_rows = rows_from_excel_com_workbook(source_path)
    else:
        raise ValueError(f"Unsupported source workbook extension: {source_path.suffix}")

    best_error: Optional[Exception] = None
    for sheet_name, rows in sheet_rows:
        if not rows:
            continue
        try:
            header_idx, cols = find_commit_source_columns(rows)
        except Exception as exc:
            best_error = exc
            continue

        log(f"[STEP] Using source sheet '{sheet_name}' with header row {header_idx + 1}")
        index: Dict[Tuple[str, str], CommitSourceData] = {}
        po_by_key: Dict[Tuple[str, str], set] = {}

        for row in rows[header_idx + 1:]:
            award_col = cols["award"]
            task_col = cols["task"]
            award = normalize_award(row[award_col] if award_col < len(row) else None)
            task = normalize_task(row[task_col] if task_col < len(row) else None)
            if not award or not task:
                continue

            key = (award, task)
            if key not in index:
                index[key] = CommitSourceData(award=award, task=task)
                po_by_key[key] = set()
            item = index[key]
            item.commitment_amount = round(item.commitment_amount + safe_float(row[cols["commitment_amount"]] if cols["commitment_amount"] < len(row) else None), 2)
            item.row_count += 1

            for optional_key in ["project_number", "project_owner_full_name", "task_name", "award_name"]:
                col = cols.get(optional_key)
                if col is not None and col < len(row):
                    val = str(row[col] or "").strip()
                    if val and not getattr(item, optional_key):
                        setattr(item, optional_key, val)

            po_col = cols.get("po_number")
            if po_col is not None and po_col < len(row):
                po_val = str(row[po_col] or "").strip()
                if po_val:
                    po_by_key[key].add(po_val)

        if index:
            for key, item in index.items():
                pos = sorted(po_by_key.get(key, set()))
                item.po_numbers = ", ".join(pos[:20])
                if len(pos) > 20:
                    item.po_numbers += f", ... +{len(pos) - 20} more"
            log(f"[STEP] Parsed {len(index)} aggregated award/task pair(s) from Commt PO Report.")
            return index
        best_error = ValueError(f"Source sheet '{sheet_name}' had headers but no parsable award/task rows.")

    if best_error:
        raise best_error
    raise ValueError("No usable Commt PO Report source sheet was found.")


def zero_source_data(award: str, task: str) -> CommitSourceData:
    return CommitSourceData(award=award, task=task, commitment_amount=0.0, row_count=0)


def resolve_shortcut(path: Path) -> Optional[Path]:
    if path.suffix.lower() != ".lnk":
        return None
    if os.name != "nt":
        log(f"[STEP] Skipping shortcut because .lnk resolution requires Windows: {path}")
        return None
    ps_script = (
        "$s=(New-Object -ComObject WScript.Shell).CreateShortcut('"
        + str(path).replace("'", "''")
        + "'); if ($s.TargetPath) { Write-Output $s.TargetPath }"
    )
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_script],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        target = result.stdout.strip()
        return Path(target) if target else None
    except Exception as exc:
        log(f"[STEP] Could not resolve shortcut {path}: {exc}")
        return None


def is_report_or_temp_file(path: Path) -> bool:
    name = path.name.lower()
    if name.startswith("~$"):
        return True
    if name.startswith("current_month_commit_fill_report"):
        return True
    if name.startswith("current_month_expense_fill_report"):
        return True
    return False


def discover_workbooks(inputs: Iterable[str], recursive: bool, source_path: Optional[Path] = None) -> List[WorkItem]:
    items: List[WorkItem] = []

    def add_file(p: Path):
        if is_report_or_temp_file(p):
            return
        if source_path is not None:
            try:
                if p.resolve() == source_path.resolve():
                    return
            except Exception:
                pass
        suffix = p.suffix.lower()
        if suffix in EXCEL_EXTS:
            items.append(WorkItem(input_path=p, workbook_path=p, source_type="file"))
        elif suffix == ".lnk":
            target = resolve_shortcut(p)
            if target and target.suffix.lower() in EXCEL_EXTS:
                items.append(WorkItem(input_path=p, workbook_path=target, source_type="shortcut"))
                log(f"[STEP] Resolved shortcut: {p} -> {target}")
            else:
                log(f"[STEP] Skipping shortcut with no Excel target: {p}")

    for raw in inputs:
        p = Path(strip_quotes(raw)).expanduser()
        if p.is_dir():
            iterator = p.rglob("*") if recursive else p.glob("*")
            for child in iterator:
                if child.is_file() and child.suffix.lower() in INPUT_EXTS:
                    add_file(child)
        elif p.is_file() and p.suffix.lower() in INPUT_EXTS:
            add_file(p)
        else:
            log(f"[STEP] Input not found or unsupported, skipping: {p}")

    seen = set()
    unique: List[WorkItem] = []
    for item in items:
        try:
            key = str(item.workbook_path.resolve()).lower()
        except Exception:
            key = str(item.workbook_path).lower()
        if key not in seen:
            seen.add(key)
            unique.append(item)
    return unique


def find_current_commit_row_openpyxl(ws) -> Optional[int]:
    for row_idx in range(1, ws.max_row + 1):
        if current_month_commit_label(ws.cell(row=row_idx, column=1).value):
            return row_idx
    return None


def find_month_column_openpyxl(ws, year: int, month: int, target_row: int) -> Tuple[Optional[int], Optional[int]]:
    max_scan_row = min(ws.max_row, max(target_row, 50))
    candidates: List[Tuple[int, int, int]] = []
    for r in range(1, max_scan_row + 1):
        for c in range(1, ws.max_column + 1):
            if month_matches(ws.cell(row=r, column=c).value, year, month):
                score = abs(target_row - r)
                if r <= target_row:
                    score -= 10000
                candidates.append((score, r, c))
    if not candidates:
        return None, None
    candidates.sort(key=lambda x: (x[0], x[1], x[2]))
    _, header_row, month_col = candidates[0]
    return header_row, month_col


def copy_style_from_prior_month_openpyxl(ws, header_row: int, target_row: int, month_col: int) -> str:
    target = ws.cell(row=target_row, column=month_col)
    for c in range(month_col - 1, 0, -1):
        if is_any_month_header(ws.cell(row=header_row, column=c).value):
            src = ws.cell(row=target_row, column=c)
            if src.has_style:
                target._style = copy(src._style)
            target.number_format = src.number_format
            target.alignment = copy(src.alignment)
            return f"{get_column_letter(c)}{target_row}"
    return ""


def process_target_openpyxl(
    item: WorkItem,
    source_index: Dict[Tuple[str, str], CommitSourceData],
    year: int,
    month: int,
) -> List[SheetResult]:
    target_path = item.workbook_path
    keep_vba = target_path.suffix.lower() == ".xlsm"
    wb = load_workbook(target_path, keep_vba=keep_vba)
    results: List[SheetResult] = []
    changed = False

    for ws in wb.worksheets:
        award = normalize_award(ws["B3"].value)
        task = normalize_task(ws["B5"].value)
        base = SheetResult(
            status="SKIP",
            input_path=str(item.input_path),
            target_workbook=str(target_path),
            sheet=ws.title,
            award=award,
            task=task,
        )

        if not award or not task:
            base.notes = "Skipped: B3 award or B5 task is blank/not usable."
            results.append(base)
            continue

        commit_row = find_current_commit_row_openpyxl(ws)
        if commit_row is None:
            base.notes = "Skipped: column A does not contain CURRENT MONTH COMMIT / CURRENT MO. COMMIT."
            results.append(base)
            continue

        header_row, month_col = find_month_column_openpyxl(ws, year, month, commit_row)
        if not header_row or not month_col:
            base.notes = "Skipped: requested month column was not found on this worksheet."
            results.append(base)
            continue

        pdata = source_index.get((award, task), zero_source_data(award, task))
        target_cell = ws.cell(row=commit_row, column=month_col)
        previous_value = target_cell.value
        style_source = copy_style_from_prior_month_openpyxl(ws, header_row, commit_row, month_col)
        target_cell.value = float(pdata.commitment_amount)
        changed = True

        diff = round(safe_float(target_cell.value) - pdata.commitment_amount, 2)
        if pdata.row_count == 0:
            status = "OK"
            notes = "No matching source rows were found for this award/task, so 0 was written."
        elif abs(diff) <= TOLERANCE:
            status = "OK"
            notes = "CURRENT MONTH COMMIT was filled from aggregated Expenditure Commitment Amount."
        else:
            status = "CHECK"
            notes = "Written value does not match the aggregated source commitment amount; please review."
        if style_source:
            notes += f" Format copied from {style_source}."

        results.append(SheetResult(
            status=status,
            input_path=str(item.input_path),
            target_workbook=str(target_path),
            sheet=ws.title,
            award=award,
            task=task,
            source_rows=pdata.row_count,
            source_commitment_amount=pdata.commitment_amount,
            written_value=pdata.commitment_amount,
            previous_value=previous_value,
            difference_written_vs_source=diff,
            current_month_commit_cell=f"{get_column_letter(month_col)}{commit_row}",
            month_header_cell=f"{get_column_letter(month_col)}{header_row}",
            style_source_cell=style_source,
            project_number=pdata.project_number,
            project_owner_full_name=pdata.project_owner_full_name,
            task_name=pdata.task_name,
            award_name=pdata.award_name,
            po_numbers=pdata.po_numbers,
            notes=notes,
        ))

    if changed:
        wb.save(target_path)
    return results


def excel_col(n: int) -> str:
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


def find_current_commit_row_com(ws, max_row: int) -> Optional[int]:
    for r in range(1, max_row + 1):
        if current_month_commit_label(ws.Cells(r, 1).Value):
            return r
    return None


def find_month_column_com(ws, year: int, month: int, target_row: int, max_row: int, max_col: int) -> Tuple[Optional[int], Optional[int]]:
    max_scan_row = min(max_row, max(target_row, 50))
    candidates: List[Tuple[int, int, int]] = []
    for r in range(1, max_scan_row + 1):
        for c in range(1, max_col + 1):
            if month_matches(ws.Cells(r, c).Value, year, month):
                score = abs(target_row - r)
                if r <= target_row:
                    score -= 10000
                candidates.append((score, r, c))
    if not candidates:
        return None, None
    candidates.sort(key=lambda x: (x[0], x[1], x[2]))
    _, header_row, month_col = candidates[0]
    return header_row, month_col


def copy_style_from_prior_month_com(excel, ws, header_row: int, target_row: int, month_col: int) -> str:
    target = ws.Cells(target_row, month_col)
    for c in range(month_col - 1, 0, -1):
        if is_any_month_header(ws.Cells(header_row, c).Value):
            ws.Cells(target_row, c).Copy()
            target.PasteSpecial(Paste=-4122)  # xlPasteFormats
            excel.CutCopyMode = False
            return f"{excel_col(c)}{target_row}"
    return ""


def process_target_excel_com(
    item: WorkItem,
    source_index: Dict[Tuple[str, str], CommitSourceData],
    year: int,
    month: int,
) -> List[SheetResult]:
    target_path = item.workbook_path
    excel, pythoncom = get_excel_app()
    wb = None
    results: List[SheetResult] = []
    changed = False
    try:
        wb = excel.Workbooks.Open(str(target_path), UpdateLinks=False, ReadOnly=False)
        for ws in wb.Worksheets:
            sheet_name = str(ws.Name)
            award = normalize_award(ws.Range("B3").Value)
            task = normalize_task(ws.Range("B5").Value)
            base = SheetResult(
                status="SKIP",
                input_path=str(item.input_path),
                target_workbook=str(target_path),
                sheet=sheet_name,
                award=award,
                task=task,
            )

            if not award or not task:
                base.notes = "Skipped: B3 award or B5 task is blank/not usable."
                results.append(base)
                continue

            used = ws.UsedRange
            max_row = used.Row + used.Rows.Count - 1
            max_col = used.Column + used.Columns.Count - 1

            commit_row = find_current_commit_row_com(ws, max_row)
            if commit_row is None:
                base.notes = "Skipped: column A does not contain CURRENT MONTH COMMIT / CURRENT MO. COMMIT."
                results.append(base)
                continue

            header_row, month_col = find_month_column_com(ws, year, month, commit_row, max_row, max_col)
            if not header_row or not month_col:
                base.notes = "Skipped: requested month column was not found on this worksheet."
                results.append(base)
                continue

            pdata = source_index.get((award, task), zero_source_data(award, task))
            target_rng = ws.Cells(commit_row, month_col)
            previous_value = target_rng.Value
            style_source = copy_style_from_prior_month_com(excel, ws, header_row, commit_row, month_col)
            target_rng.Value = float(pdata.commitment_amount)
            changed = True
            try:
                excel.Calculate()
            except Exception:
                pass

            diff = round(safe_float(target_rng.Value) - pdata.commitment_amount, 2)
            if pdata.row_count == 0:
                status = "OK"
                notes = "No matching source rows were found for this award/task, so 0 was written."
            elif abs(diff) <= TOLERANCE:
                status = "OK"
                notes = "CURRENT MONTH COMMIT was filled from aggregated Expenditure Commitment Amount."
            else:
                status = "CHECK"
                notes = "Written value does not match the aggregated source commitment amount; please review."
            if style_source:
                notes += f" Format copied from {style_source}."

            results.append(SheetResult(
                status=status,
                input_path=str(item.input_path),
                target_workbook=str(target_path),
                sheet=sheet_name,
                award=award,
                task=task,
                source_rows=pdata.row_count,
                source_commitment_amount=pdata.commitment_amount,
                written_value=pdata.commitment_amount,
                previous_value=previous_value,
                difference_written_vs_source=diff,
                current_month_commit_cell=f"{excel_col(month_col)}{commit_row}",
                month_header_cell=f"{excel_col(month_col)}{header_row}",
                style_source_cell=style_source,
                project_number=pdata.project_number,
                project_owner_full_name=pdata.project_owner_full_name,
                task_name=pdata.task_name,
                award_name=pdata.award_name,
                po_numbers=pdata.po_numbers,
                notes=notes,
            ))

        if changed:
            wb.Save()
        return results
    finally:
        if wb is not None:
            wb.Close(SaveChanges=False)
        excel.Quit()
        pythoncom.CoUninitialize()


def write_processing_report(output_dir: Path, results: List[SheetResult], month_key: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = output_dir / f"current_month_commit_fill_report_{timestamp}.xlsx"

    headers = [
        "status",
        "input_path",
        "target_workbook",
        "sheet",
        "award",
        "task",
        "source_rows",
        "source_expenditure_commitment_amount",
        "written_current_month_commit",
        "previous_current_month_commit",
        "difference_written_vs_source",
        "current_month_commit_cell",
        "month_header_cell",
        "style_source_cell",
        "month",
        "project_number",
        "project_owner_full_name",
        "task_name",
        "award_name",
        "po_numbers",
        "notes",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Processing Report"
    ws.append(headers)

    for r in results:
        ws.append([
            r.status,
            r.input_path,
            r.target_workbook,
            r.sheet,
            r.award,
            r.task,
            r.source_rows,
            r.source_commitment_amount,
            r.written_value,
            r.previous_value,
            r.difference_written_vs_source,
            r.current_month_commit_cell,
            r.month_header_cell,
            r.style_source_cell,
            month_key,
            r.project_number,
            r.project_owner_full_name,
            r.task_name,
            r.award_name,
            r.po_numbers,
            r.notes,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    status_col = headers.index("status") + 1
    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row_idx, status_col).value or "").upper()
        if status == "OK":
            fill = PatternFill("solid", fgColor="E2F0D9")
        elif status == "CHECK":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif status == "SKIP":
            fill = PatternFill("solid", fgColor="E7E6E6")
        else:
            fill = PatternFill("solid", fgColor="FCE4D6")
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).fill = fill

    currency_headers = {
        "source_expenditure_commitment_amount",
        "written_current_month_commit",
        "previous_current_month_commit",
        "difference_written_vs_source",
    }
    for col_idx, header in enumerate(headers, 1):
        if header in currency_headers:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = '#,##0.00;[Red](#,##0.00);-'
        max_len = len(header)
        for row_idx in range(2, min(ws.max_row, 200) + 1):
            max_len = max(max_len, len(str(ws.cell(row_idx, col_idx).value or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 70)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(report_path)
    return report_path


def choose_output_dir(items: List[WorkItem], output_dir_arg: Optional[str]) -> Path:
    if output_dir_arg:
        return Path(strip_quotes(output_dir_arg))
    if items:
        # If input is a shortcut, report defaults to the shortcut folder, not the target folder.
        return items[0].input_path.parent
    return Path.cwd()


def ask_path(prompt: str) -> str:
    return strip_quotes(input(prompt).strip())


def process_all_targets(
    items: List[WorkItem],
    source_index: Dict[Tuple[str, str], CommitSourceData],
    year: int,
    month: int,
    use_excel_com: bool,
) -> List[SheetResult]:
    all_results: List[SheetResult] = []
    for item in items:
        path = item.workbook_path
        log(f"\n[FILE] {path}")
        try:
            if path.suffix.lower() == ".xls" or use_excel_com:
                if not excel_com_available():
                    raise RuntimeError("Microsoft Excel COM is required for .xls targets or --use-excel-com, but it is not available.")
                results = process_target_excel_com(item, source_index, year, month)
            else:
                results = process_target_openpyxl(item, source_index, year, month)
            all_results.extend(results)

            updated = sum(1 for r in results if r.written_value is not None)
            check_count = sum(1 for r in results if r.status == "CHECK")
            skip_count = sum(1 for r in results if r.status == "SKIP")
            print(f"[STEP] {path.name}: updated {updated}, check {check_count}, skipped {skip_count}")
        except Exception as exc:
            all_results.append(SheetResult(
                status="FAIL",
                input_path=str(item.input_path),
                target_workbook=str(path),
                sheet="",
                notes=str(exc),
            ))
            print(f"[FAIL] {path}: {exc}")
    return all_results


def main() -> int:
    parser = argparse.ArgumentParser(description="Fill CURRENT MONTH COMMIT by worksheet/tab from Commt PO Report source data.")
    parser.add_argument("--source", help="Path to the Commt PO Report workbook.")
    parser.add_argument("--targets", nargs="+", help="Target workbook(s), folder(s), or .lnk shortcut(s) to process.")
    parser.add_argument("--month", help="Target workbook month column to fill, e.g. apr-2026.")
    parser.add_argument("--output-dir", help="Folder where the processing report should be written. Defaults to first target input folder.")
    parser.add_argument("--include-subfolders", action="store_true", help="When a target input is a folder, scan subfolders too.")
    parser.add_argument("--use-excel-com", action="store_true", help="Use Microsoft Excel COM for all target workbooks. Required automatically for .xls targets.")
    args = parser.parse_args()

    if not args.source:
        args.source = ask_path("Enter FULL path to the Commt PO Report source workbook: ")
    if not args.month:
        args.month = strip_quotes(input("Enter target workbook month column to fill, e.g. apr-2026: "))
    if not args.targets:
        args.targets = [ask_path("Enter FULL path to target workbook, target folder, or .lnk shortcut: ")]

    source_path = Path(strip_quotes(args.source)).expanduser()
    target_inputs = [strip_quotes(x) for x in args.targets]

    print("\n=== RUNNING WITH ===")
    print(f"Source workbook: {source_path}")
    print(f"Targets:         {target_inputs}")
    print(f"Month:           {args.month}")
    print(f"Include subfolders: {args.include_subfolders}")
    print("Backup files:    NO BACKUP WILL BE CREATED")
    print("====================\n")

    if not source_path.exists():
        print(f"[FAIL] Source workbook not found: {source_path}")
        return 1
    if source_path.suffix.lower() not in EXCEL_EXTS:
        print(f"[FAIL] Unsupported source workbook type: {source_path.suffix}")
        return 1

    try:
        year, month, month_key = parse_month_input(args.month)
    except Exception as exc:
        print(f"[FAIL] Month could not be parsed: {exc}")
        return 1

    try:
        source_index = read_commit_source_data(source_path)
    except Exception as exc:
        print(f"[FAIL] Source workbook could not be read: {exc}")
        return 1

    items = discover_workbooks(target_inputs, recursive=args.include_subfolders, source_path=source_path)
    print(f"[STEP] Found {len(items)} target workbook(s)/shortcut target(s) to process.")
    if not items:
        print("[FAIL] No target Excel files were found from the provided target input.")
        return 1

    output_dir = choose_output_dir(items, args.output_dir)
    all_results = process_all_targets(items, source_index, year, month, args.use_excel_com)

    try:
        report_path = write_processing_report(output_dir, all_results, month_key)
        print(f"\n[STEP] Processing report written to: {report_path}")
    except Exception as exc:
        print(f"[FAIL] Could not write processing report: {exc}")
        return 1

    updated = sum(1 for r in all_results if r.written_value is not None)
    ok_count = sum(1 for r in all_results if r.status == "OK")
    check_count = sum(1 for r in all_results if r.status == "CHECK")
    skip_count = sum(1 for r in all_results if r.status == "SKIP")
    fail_count = sum(1 for r in all_results if r.status == "FAIL")

    print("\n=== SUMMARY ===")
    print(f"Target files:    {len(items)}")
    print(f"Sheets updated:  {updated}")
    print(f"OK:              {ok_count}")
    print(f"CHECK:           {check_count}")
    print(f"SKIP:            {skip_count}")
    print(f"FAIL:            {fail_count}")

    if check_count:
        print("\nItems needing review because written value did not match the aggregated source commitment amount:")
        for r in all_results:
            if r.status == "CHECK":
                print(f"  - {r.target_workbook} | {r.sheet}: {r.award}/{r.task} | written {r.written_value} vs source {r.source_commitment_amount}")

    if fail_count:
        return 2
    if check_count:
        return 3
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
