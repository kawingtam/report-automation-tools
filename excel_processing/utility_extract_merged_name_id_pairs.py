from openpyxl import load_workbook
import os
import csv

INPUT_FOLDER = input("Enter input folder path: ").strip().strip('"')
OUTPUT_CSV = input("Enter output CSV file path: ").strip().strip('"')

def get_merged_cell_value(ws, row, col):
    cell = ws.cell(row=row, column=col)

    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(merged_range.min_row, merged_range.min_col).value

    return None


def extract_from_file(filepath):
    wb = load_workbook(filepath, data_only=True)
    results = []

    for ws in wb.worksheets:
        seen = set()  # 🔥 prevents duplicates from merged cells

        for row in range(11, ws.max_row + 1):  # ✅ start at row 11
            name_value = get_merged_cell_value(ws, row, 3)   # Column C
            number_value = get_merged_cell_value(ws, row, 4) # Column D

            # ✅ only keep rows where BOTH exist
            if name_value and number_value:
                key = (str(name_value).strip(), str(number_value).strip())

                if key not in seen:
                    seen.add(key)

                    results.append({
                        "file_name": os.path.basename(filepath),
                        "sheet_name": ws.title,
                        "name": str(name_value).strip(),
                        "number": str(number_value).strip()
                    })

    return results


def main():
    all_results = []

    for file_name in os.listdir(INPUT_FOLDER):
        if file_name.endswith(".xlsx"):
            full_path = os.path.join(INPUT_FOLDER, file_name)
            print(f"Processing: {file_name}")
            file_results = extract_from_file(full_path)
            all_results.extend(file_results)

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["file_name", "sheet_name", "name", "number"]
        )
        writer.writeheader()
        writer.writerows(all_results)

    print(f"\n✅ Done. Extracted {len(all_results)} unique rows to:")
    print(OUTPUT_CSV)


if __name__ == "__main__":
    main()