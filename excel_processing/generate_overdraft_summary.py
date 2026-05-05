"""
Overdraft Summary Report Generator

Reads an Excel pivot-table style worksheet and creates an output workbook named:
OD_SUMMARY_YYYY-MM-DD_HHMMSS.xlsx
in the same folder as the source workbook.

Expected pivot layout:
- One column named Row Labels / Row Label / Labels contains awards and tasks.
- Award rows are usually text codes like ABCDE, FGHIJ.
- Task rows are usually numeric labels like 601, 603, 4961.
- The OD amount is detected from the monthly Actual Expenditure Amount column,
  for example: Sum of MAR-2026 Actual Expenditure Amount.

The script lists every award-task row where that amount is negative:
ABCDE-601 There is an OD ($1,904.53).
"""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from typing import Any, Optional

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: The Python package 'openpyxl' is required.")
    print("Install it with: py -m pip install openpyxl")
    input("Press Enter to exit...")
    sys.exit(1)


TASK_RE = re.compile(r"^\d+(?:\.0)?$")
AWARD_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_-]*$")


def clean_text(value: Any) -> str:
    """Convert Excel cell value to a clean string."""
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and TASK_RE.match(text):
        text = text[:-2]
    return text


def to_number(value: Any) -> Optional[float]:
    """Convert Excel values like -1904.53 or ($1,904.53) to float."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1]

    text = text.replace("$", "").replace(",", "").replace(" ", "")
    try:
        num = float(text)
        return -num if negative else num
    except ValueError:
        return None


def format_currency(amount: float) -> str:
    """Format negative OD amount as ($1,904.53)."""
    amount = abs(float(amount))
    return f"(${amount:,.2f})"


def find_pivot_header(ws) -> tuple[int, int, int]:
    """
    Find the header row, row-label column, and monthly actual amount column.
    Returns: header_row, label_col, od_col
    """
    best_candidates = []

    for row in range(1, min(ws.max_row, 30) + 1):
        headers = [clean_text(ws.cell(row=row, column=col).value) for col in range(1, ws.max_column + 1)]

        label_col = None
        for idx, header in enumerate(headers, start=1):
            h = header.lower()
            if h in {"row labels", "row label", "labels", "label"}:
                label_col = idx
                break

        if not label_col:
            continue

        # Prefer monthly actual expenditure column, excluding FTD/PTD cumulative columns.
        for idx, header in enumerate(headers, start=1):
            h = header.lower()
            if "actual expenditure amount" in h and "ftd" not in h and "ptd" not in h:
                best_candidates.append((row, label_col, idx, 1))

        # Fallback if only a generic actual expenditure column exists.
        for idx, header in enumerate(headers, start=1):
            h = header.lower()
            if "actual expenditure" in h and "ftd" not in h and "ptd" not in h:
                best_candidates.append((row, label_col, idx, 2))

    if not best_candidates:
        raise ValueError(
            "Could not find the pivot headers. Make sure the pivot table has a 'Row Labels' column "
            "and a monthly 'Actual Expenditure Amount' column."
        )

    best_candidates.sort(key=lambda x: x[3])
    header_row, label_col, od_col, _ = best_candidates[0]
    return header_row, label_col, od_col


def find_pivot_sheet(wb) -> tuple[Any, int, int, int]:
    """Find the first worksheet that contains the pivot table headers."""
    errors = []
    for ws in wb.worksheets:
        try:
            header_row, label_col, od_col = find_pivot_header(ws)
            return ws, header_row, label_col, od_col
        except ValueError as exc:
            errors.append(f"{ws.title}: {exc}")
    raise ValueError("No pivot table found in this workbook.\n" + "\n".join(errors))


def is_award_label(label: str) -> bool:
    """Award labels are usually alphabetic/alphanumeric codes, not task numbers."""
    return bool(label) and not TASK_RE.match(label) and AWARD_RE.match(label)


def is_task_label(label: str) -> bool:
    """Task labels are usually numeric codes."""
    return bool(TASK_RE.match(label))


def build_summary_rows(ws, header_row: int, label_col: int, od_col: int) -> list[list[Any]]:
    """Extract all award-task rows with negative monthly actual amount."""
    current_award = ""
    results = []

    for row in range(header_row + 1, ws.max_row + 1):
        label = clean_text(ws.cell(row=row, column=label_col).value)
        if not label:
            continue

        lower_label = label.lower()
        if lower_label.startswith("grand total") or lower_label == "total":
            continue

        if is_award_label(label):
            current_award = label.upper()
            continue

        if not current_award or not is_task_label(label):
            continue

        amount = to_number(ws.cell(row=row, column=od_col).value)
        if amount is not None and amount < 0:
            award_task = f"{current_award}-{label}"
            sentence = f"{award_task} There is an OD {format_currency(amount)}."
            results.append([current_award, label, amount, sentence])

    return results


def autosize_columns(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 70)


def create_output_workbook(source_path: str, source_sheet: str, od_header: str, rows: list[list[Any]]) -> str:
    source_dir = os.path.dirname(os.path.abspath(source_path))
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    output_path = os.path.join(source_dir, f"OD_SUMMARY_{timestamp}.xlsx")

    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "PTA Summary"

    title = "Overdraft Summary"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:D1")

    ws["A2"] = "Source file"
    ws["B2"] = os.path.basename(source_path)
    ws["A3"] = "Source sheet"
    ws["B3"] = source_sheet
    ws["A4"] = "OD column used"
    ws["B4"] = od_header
    ws["A5"] = "Generated"
    ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    headers = ["Award", "Task", "OD Amount", "Summary"]
    ws.append([])
    ws.append(headers)
    header_row = 7

    for row_data in rows:
        ws.append(row_data)

    # Style title metadata.
    meta_bold = Font(bold=True)
    for r in range(2, 6):
        ws.cell(r, 1).font = meta_bold

    # Style table.
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[2].number_format = '($#,##0.00);($#,##0.00);$0.00'

    if not rows:
        ws.append(["", "", "", "No overdraft rows found."])

    ws.freeze_panes = "A8"
    autosize_columns(ws)
    ws.column_dimensions["D"].width = 55

    out_wb.save(output_path)
    return output_path


def main() -> None:
    print("=== PTA SUMMARY OD REPORT TOOL ===")

    if len(sys.argv) >= 2:
        input_path = sys.argv[1].strip().strip('"')
    else:
        input_path = input("Enter FULL path to Excel pivot table: ").strip().strip('"')

    if not input_path:
        print("No file path entered.")
        input("Press Enter to exit...")
        return

    if not os.path.exists(input_path):
        print(f"ERROR: File not found: {input_path}")
        input("Press Enter to exit...")
        return

    try:
        wb = load_workbook(input_path, data_only=True)
        ws, header_row, label_col, od_col = find_pivot_sheet(wb)
        od_header = clean_text(ws.cell(row=header_row, column=od_col).value)
        rows = build_summary_rows(ws, header_row, label_col, od_col)
        output_path = create_output_workbook(input_path, ws.title, od_header, rows)

        print(f"\nDone. Found {len(rows)} overdraft award-task row(s).")
        print(f"Output saved here:\n{output_path}")
    except Exception as exc:
        print("\nERROR:")
        print(exc)

    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
